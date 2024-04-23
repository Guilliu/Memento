import numpy as np, pandas as pd, statsmodels.api as sm, datetime

from scipy import special
from scipy.stats import ks_2samp
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import train_test_split
from sklearn.tree import _tree, DecisionTreeClassifier


####################################################################################################


N = 120


def string_categories1(x, y):

    return dict(map(reversed, enumerate(pd.Series(y).groupby(
    x, dropna=False).mean().sort_values(ascending=False).index.values)))


def string_categories2(breakpoints_cat):

    try: return dict(map(reversed, enumerate([i for j in breakpoints_cat for i in j])))
    except: return {}


def string_to_num(x, categories):

    return pd.Series(x).map(categories).values


def breakpoints_to_str(breakpoints_num, categories):
 
    breakpoints_str = []

    for i in range(len(breakpoints_num)):
        if i == 0:
            breakpoints_str.append([j[0] for j in categories.items()
            if j[1] < breakpoints_num[i]])
        else:
            breakpoints_str.append([j[0] for j in categories.items()
            if breakpoints_num[i-1] <= j[1] < breakpoints_num[i]])
        if i == len(breakpoints_num) - 1:
            breakpoints_str.append([j[0] for j in categories.items()
            if breakpoints_num[i] <= j[1]])

    return breakpoints_str


def breakpoints_to_num(breakpoints_cat):

    if isinstance(breakpoints_cat[0], list):

        L, suma = [], 0
        for i in breakpoints_cat[:-1]:
            suma += len(i)
            L.append(suma-0.5)
        return np.array(L)

    else: return breakpoints_cat


def remapeo_missing(v, bp, old_value=-12345678):

    if isinstance(bp, dict):

        breakpoints = bp['bp']
        missing_group = bp['mg']
        if missing_group != None:
            if missing_group!= 0:
                if missing_group == 1:
                    return np.where(v == old_value, breakpoints[missing_group-1]-(np.e-2), v)
                if missing_group >= 2:
                    return np.where(v == old_value, breakpoints[missing_group-2]+(np.e-2), v)
            else: return v
        else: return v
    else: return v


def data_convert(x, categories):

    x_original = x

    if x.dtype in ('O', 'bool'):
        if categories == {}:
            raise ValueError('En una variable de tipo texto o booleana '
            'es necesario especificar el diccionario de categorias')
        if pd.Series(x).isna().sum() > 0:
            x = pd.Series(x).replace(np.nan, 'Missing').values
        x_initial = x
        x = string_to_num(x, categories)
        x_converted = x

    else:
        x_initial = x
        x_converted = x

    if x.dtype not in ('O', 'bool') and np.isnan(x).sum() > 0:
        x_final = np.nan_to_num(x, nan=-12345678)

    else: x_final = x_converted

    return x_original, x_initial, x_converted, x_final


def adapt_data(X, y, variables, breakpoints, target_name='target_4815162342'):
    
    df = pd.DataFrame()
    for variable in variables:
        
            bp = breakpoints[variable]
            if isinstance(bp, dict) and bp['mg'] == None: bp = bp['bp']
            
            if not isinstance(bp, dict):
                df[variable] = data_convert(X[variable].values, string_categories2(bp))[3]
                
            else:
                df[variable] = remapeo_missing(data_convert(
                X[variable].values, string_categories2(bp))[3], bp)
        
    df[target_name] = y

    return df


def split(X, y, test_seed=123, test_size=0.3, stratify=True,
stratify_var='', flag_train_test=[], verbose=False):
    
    if flag_train_test != []:

        try: a, b, c = flag_train_test
        except:
            print('En la variable flag_train_test hay que introducir tres cosas en orden: '
            'el nombre de la variable con el flag, el valor de train y el valor de test')

        data = X.copy()
        data['target_4815162342'] = y

        X_train = data[data[a] == b].drop('target_4815162342', axis=1)
        y_train = data[data[a] == b]['target_4815162342'].values

        X_test = data[data[a] == c].drop('target_4815162342', axis=1)
        y_test = data[data[a] == c]['target_4815162342'].values

    else:

        if stratify:

            if stratify_var == '':
                X_train, X_test, y_train, y_test = train_test_split(X, y,
                test_size=test_size, random_state=test_seed, stratify=y)

            else:
                X_train, X_test, y_train, y_test = train_test_split(X, y,
                test_size=test_size, random_state=test_seed,
                stratify=X[stratify_var])

        else:
            X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=test_size, random_state=test_seed)

    index_train, index_test = X_train.index, X_test.index
    X_train, X_test = X_train.reset_index(drop=True), X_test.reset_index(drop=True)

    if isinstance(y_train, pd.Series): y_train, y_test = y_train.values, y_test.values
    y_train, y_test = y_train, y_test

    if verbose == True and flag_train_test == []:

        if stratify:

            if stratify_var == '':
                print('Particionado {}-{} estratificado en el target terminado'\
                .format(int(100*(1-test_size)), int(100*test_size)))

            else:
                print('Particionado {}-{} estratificado en la variable \'{}\' terminado'\
                .format(int(100*(1-test_size)), int(100*test_size), stratify_var))

        else:
            print('Particionado {}-{} terminado'.format(int(100*(1-test_size)), int(100*test_size)))

        print('-' * N)
        
    return X_train, X_test, y_train, y_test, index_train, index_test


def autogrouping(X, y, variables=[], autogrp_max_groups=5, autogrp_min_pct=0.05,
autogrp_dict_max_groups={}, autogrp_dict_min_pct={}, autogrp_dict_manual_types={}, verbose=False):
    
    autogroupings = {}
    variables_no_agrupadas_error = []
    
    if variables == []: variables = X.columns  
    for variable in variables:

        try:

            if variable in autogrp_dict_max_groups:
                max_groups = autogrp_dict_max_groups[variable]
            else: max_groups = autogrp_max_groups

            if variable in autogrp_dict_min_pct:
                min_pct = autogrp_dict_min_pct[variable]
            else: min_pct = autogrp_min_pct

            if variable in autogrp_dict_manual_types: 
                manual_type = autogrp_dict_manual_types[variable]
            else: manual_type = ''

            x = X[variable].values
            frenken = Autogrouping(max_groups=max_groups, 
            min_pct=min_pct, manual_type=manual_type).fit(x, y)

            if len(frenken.breakpoints_num) == 0: variables_no_agrupadas_error.append(variable)
            else: autogroupings[variable] = frenken

        except:
            variables_no_agrupadas_error.append(variable)
            
    if verbose:
        
        print('Autogrouping terminado. Máximo número de buckets = {}. Mínimo porcentaje '
        'por bucket = {}'.format(autogrp_max_groups, autogrp_min_pct))
        print('-' * N)

        if len(variables_no_agrupadas_error) > 0:
            print('Variables no agrupadas por algún error, seguramente por excesiva '
            'concentración en algún valor (> 95%) : {}'.format(variables_no_agrupadas_error))
            print('-' * N)
        
    return autogroupings, variables_no_agrupadas_error


def compute_group_names(dtype, breakpoints, missing_group=None, decimals=2):

    if dtype in ('O', 'bool'): return breakpoints

    else:
        
        if isinstance(breakpoints, dict): breakpoints = breakpoints['bp']

        groups = np.concatenate([[-np.inf], breakpoints, [np.inf]])
        group_names1, group_names2 = [], []

        for i in range(len(groups) - 1):

            if np.isinf(groups[i]):
                a = '({0:.{2}f}, {1:.{2}f})'\
                .format(groups[i], groups[i+1], decimals)
            else: a = '[{0:.{2}f}, {1:.{2}f})'.format(groups[i], groups[i+1], decimals)
            group_names1.append(a)

        for group in group_names1:
            if '-12345670.00)' in group: group = 'Missing'
            if '[-12345670.00' in group: group = group.replace('[-12345670.00', '(-inf')
            group_names2.append(group)

        if isinstance(missing_group, int):
            if missing_group > 0:
                group_names2[missing_group-1] += ', Missing'

        return group_names2


def compute_table(x, y, breakpoints_num, group_names, compute_totals=True):

    x_groups = np.digitize(x, breakpoints_num)
 
    ngroups = len(breakpoints_num) + 1
    g = np.zeros(ngroups).astype(np.int64)
    b = np.zeros(ngroups).astype(np.int64)

    for i in range(ngroups):

        g[i] = np.sum([(y == 0) & (x_groups == i)])
        b[i] = np.sum([(y == 1) & (x_groups == i)])

    e = g + b

    total_g = g.sum()
    total_b = b.sum()
    total_e = e.sum()

    pct_g = g / total_g
    pct_b = b / total_b
    pct_e = e / total_e

    b_rate = b / e
    woe = np.log(1 / b_rate - 1) + np.log(total_b / total_g)
    iv = special.xlogy(pct_b - pct_g, pct_b / pct_g)

    total_b_rate = total_b / total_e
    total_iv = iv.sum()

    table = pd.DataFrame({

        'Group': group_names,
        'Count': e,
        'Percent': pct_e,
        'Goods': g,
        'Bads': b,
        'Bad rate': b_rate,
        'WoE': woe,
        'IV': iv,
    })

    if compute_totals:
        table.loc['Totals'] = ['', total_e, 1, total_g, total_b, total_b_rate, '', total_iv]

    return table, total_iv


def transform_to_woes(x, y, breakpoints_num):

    x_groups = np.digitize(x, breakpoints_num)

    ngroups = len(breakpoints_num) + 1
    g = np.empty(ngroups).astype(np.int64)
    b = np.empty(ngroups).astype(np.int64)

    for i in range(ngroups):

        g[i] = np.sum([(y == 0) & (x_groups == i)])
        b[i] = np.sum([(y == 1) & (x_groups == i)])

    e = g + b

    total_g = g.sum()
    total_b = b.sum()
    b_rate = b / e

    woe = np.log(1 / b_rate - 1) + np.log(total_b / total_g)

    mapeo_indices_woes = dict(zip([i for i in range(len(woe))], list(woe)))
    x_woes = pd.Series([mapeo_indices_woes[i] for i in x_groups])

    return x_woes


def calib_score(points, num_variables, intercept, calibracion='default'):

    n = num_variables
    
    if calibracion == 'default': pdo, odds, scorecard_points = 20, 1, 500
    else: pdo, odds, scorecard_points = calibracion

    factor = pdo / np.log(2)
    offset = scorecard_points - factor * np.log(odds)

    new_points = -(points + intercept / n) * factor + offset / n

    return new_points


def compute_scorecard(data, features, info, target_name='target_4815162342',
pvalues=False, ret_coefs=False, redondeo=True, logistic_method='newton', calibracion='default'):

    X = data.drop(target_name, axis=1).copy()
    y = data[target_name].values

    Xwoes = pd.DataFrame()
    scorecard, features_length = pd.DataFrame(), np.array([], 'int64')

    for feature in features:

        x = X[feature].values
        breakpoints_num = info[feature]['breakpoints_num']
        group_names = info[feature]['group_names']

        table = compute_table(x, y, breakpoints_num, group_names, False)[0]
        table.insert(0, 'Variable', feature)
        scorecard = pd.concat([scorecard, table])
        features_length = np.append(features_length, len(table))

        Xwoes[feature] = transform_to_woes(x, y, breakpoints_num)

    log_reg = sm.Logit(y, sm.add_constant(Xwoes.values)).fit(method=logistic_method, disp=0)
    coefs, intercept = np.array([log_reg.params[1:]]), np.array([log_reg.params[0]])

    scorecard['Raw score'] = scorecard['WoE'] * np.repeat(coefs.ravel(), features_length)
    scorecard['Aligned score'] = calib_score(scorecard['Raw score'], len(features), intercept, calibracion)
    if redondeo: scorecard['Aligned score'] = scorecard['Aligned score'].round().astype('int')
    scorecard = scorecard.reset_index(drop=True)

    if pvalues and ret_coefs: return scorecard, features_length, log_reg.pvalues, coefs.ravel()
    if pvalues and not ret_coefs: return scorecard, features_length, log_reg.pvalues
    if not pvalues and ret_coefs: return scorecard, features_length, coefs.ravel()
    if not pvalues and not ret_coefs: return scorecard, features_length


def transform_to_points(x, breakpoints_num, mapeo_points):

    return pd.Series([mapeo_points[i] for i in np.digitize(x, breakpoints_num)])


def apply_scorecard(data, scorecard, info, target_name='',
binary_treshold=0.0, score_name='scorecardpoints'):

    features = list(scorecard['Variable'].unique())
    if target_name != '': columnas = features + [target_name]
    else: columnas = features
                 
    data_final = data[columnas].copy()
    data_final = data.copy()
    data_final[score_name] = 0

    for feature in features:

        x = data_final[feature]
        breakpoints_num = info[feature]['breakpoints_num']

        mapeo_points = scorecard[scorecard['Variable'] == feature]\
        .reset_index(drop=True)['Aligned score'].to_dict()

        data_final['scr_{}'.format(feature)] = \
        transform_to_points(x, breakpoints_num, mapeo_points)
        data_final[score_name] += data_final['scr_{}'.format(feature)]
    
    columnas = list(data_final.columns)
    columnas.remove(score_name)
    data_final = data_final[columnas + [score_name]]

    if binary_treshold != 0.0:
        data_final['prediction'] = np.where(data_final[score_name] >= binary_treshold, 0, 1)

    return data_final

def compute_metrics(data, target_name, metrics, print_log=False, score_name='scorecardpoints'):

    if metrics not in (['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']):
        raise ValueError("Valor erroneo para 'metrics'. Los valores "
        "váidos son: ['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']")

    if 'ks' in metrics:
        g = data.loc[data[target_name] == 0, score_name]
        b = data.loc[data[target_name] == 1, score_name]
        ks = ks_2samp(g, b)[0]

    if 'gini' in metrics:
        gini = abs(2*(1 - roc_auc_score(data[target_name], data[score_name])) - 1)

    if metrics == ['ks']:
        if print_log:
            print('El modelo tiene un {:.2f}% de KS en esta muestra'.format(round(ks*100, 2)))
        return ks

    if metrics == ['gini']:
        if print_log:
            print('El modelo tiene un {:.2f}% de Gini en esta muestra'.format(round(gini*100, 2), ))
        return gini

    if metrics in (['ks', 'gini'], ['gini', 'ks']):
        if print_log:
            print('El  modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini '
            'en esta muestra'.format(round(ks*100, 2), round(gini*100, 2)))
        return ks, gini


def compute_final_breakpoints(variables, autogroupings, user_breakpoints):

    final_breakpoints = {}

    for variable in variables:
        try: final_breakpoints[variable] = user_breakpoints[variable]
        except: final_breakpoints[variable] = autogroupings[variable].breakpoints

    return final_breakpoints


def compute_info(X, variables, breakpoints):
    
    info = {}
    for variable in variables:
        
            info[variable] = {}
            bp = breakpoints[variable]
            
            if not isinstance(bp, dict):
                info[variable]['breakpoints_num'] = breakpoints_to_num(bp)
                info[variable]['group_names'] = compute_group_names(X[variable].values.dtype, bp)
                
            else:
                info[variable]['breakpoints_num'] = breakpoints_to_num(bp['bp'])
                info[variable]['group_names'] = compute_group_names(
                X[variable].values.dtype, bp['bp'], bp['mg'])
                
    return info


def features_selection(data, features, var_list, info, target_name='target_4815162342',
method='stepwise', metric='pvalue', threshold=0.01, stop_ks_gini=True, 
max_iters=14, included_vars=[], muestra_test=None, show='gini',
check_overfitting=True, logistic_method='newton', calibracion='default'):
    
    # if log_file: file_prints = open('log_modelo.txt', 'a')
    # else: file_prints = None

    if features != []: included_vars, max_iters = features, 0

    if method not in ('forward', 'stepwise'):
        raise ValueError("Valor inválido para el parámetro 'method', "
        "solo están pertimidos los valores 'forward' y 'stepwise'")

    if metric not in ('pvalue', 'ks', 'gini'):
        raise ValueError("Valor inválido para el parámetro 'metric', "
        "solo están pertimidos los valores 'pvalue', 'ks' y 'gini")


    if max_iters > len(var_list):
        print('Cuidado, has puesto un valor numero máximo de iteraciones ({})'
        ' superior al número de variables candidatas ({})'.format(max_iters, len(var_list)))
        print('-' * N)
        max_iters = len(var_list)
    
    if check_overfitting:
        
        posible_sobreajuste = []
        for var in var_list:
            
            try:
                scorecard, features_length = compute_scorecard(
                data, [var], info, target_name=target_name, logistic_method=logistic_method, calibracion=calibracion)
                data_final = apply_scorecard(data, scorecard, info, target_name)
                gini_train = compute_metrics(data_final, target_name, ['gini'])
                if not isinstance(muestra_test, type(None)):
                    test_final = apply_scorecard(muestra_test, scorecard, info, target_name)
                    gini_test = compute_metrics(test_final, target_name, ['gini'])
                if gini_test + 0.25 < gini_train:
                    posible_sobreajuste.append(var)
                    
            except: pass
        
        if len(posible_sobreajuste) > 0:
            print('Posibles variables con overfitting: {}'.format(posible_sobreajuste))
            print('-' * N)
            
    old_ks, old_gini = 0, 0
    
    features = []
    
    num_included = len(included_vars)
    for i in range(num_included + max_iters):

        if i < num_included:

            new_var = included_vars.pop(0)
            features.append(new_var)

            if metric == 'pvalue':

                scorecard, features_length, pvalues = compute_scorecard(
                data, features, info, target_name=target_name, pvalues=True, logistic_method=logistic_method, calibracion=calibracion)
                train_final = apply_scorecard(data, scorecard, info, target_name)
                ks_train, gini_train = compute_metrics(train_final, target_name, ['ks', 'gini'])
                if not isinstance(muestra_test, type(None)):
                    test_final = apply_scorecard(muestra_test, scorecard, info, target_name)
                    ks_test, gini_test = compute_metrics(test_final, target_name, ['ks', 'gini'])
                
                if pvalues[-1] < 1e-100: pvalorcito = 0
                else: pvalorcito = pvalues[-1]

                if isinstance(muestra_test, type(None)):
                    if show == 'ks':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} '
                        '| KS train = {:.2f}% ---> Feature selected: {}'
                        .format(str(i+1).zfill(2), pvalorcito, ks_train*100, var))
                    if show == 'gini':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} '
                        '| Gini train = {:.2f}% ---> Feature selected: {}'
                        .format(str(i+1).zfill(2), pvalorcito, gini_train*100, var))
                    if show == 'both':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} '
                        '| KS train = {:.2f}% | Gini train = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), pvalorcito, 
                        ks_train*100, gini_train*100, new_var))

                else:
                    if show == 'ks':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} '
                        '| KS train = {:.2f}% | KS test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), pvalorcito,
                        ks_train*100, ks_test*100, new_var))
                    if show == 'gini':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} '\
                        '| Gini train = {:.2f}% | Gini test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), pvalorcito,
                        gini_train*100, gini_test*100, new_var))
                    if show == 'both':
                        print('Step {} | 0:00:00.000000 | pv = {:.2e} | KS train = {:.2f}% '
                        '| KS test = {:.2f}% | Gini train = {:.2f}% | Gini test '
                        '= {:.2f}% ---> Feature selected: {}'.format(str(i+1).zfill(2), pvalorcito,
                         ks_train*100, ks_test*100, gini_train*100, gini_test*100, new_var)) 

            else:

                scorecard, features_length = compute_scorecard(
                data, features, info, target_name=target_name, logistic_method=logistic_method, calibracion=calibracion)
                train_final = apply_scorecard(data, scorecard, info, target_name)
                ks_train, gini_train = compute_metrics(train_final, target_name, ['ks', 'gini'])
                if not isinstance(muestra_test, type(None)):
                    test_final = apply_scorecard(muestra_test, scorecard, info, target_name)
                    ks_test, gini_test = compute_metrics(test_final, target_name, ['ks', 'gini'])

                if isinstance(muestra_test, type(None)):
                    if show == 'ks':
                        print('Step {} | 0:00:00.000000 | KS train = {:.2f}% '
                        '---> Feature selected: {}'.format(
                        str(i+1).zfill(2), ks_train*100, new_var),)
                    if show == 'gini':
                        print('Step {} | 0:00:00.000000 | Gini train = {:.2f}% '
                        '---> Feature selected: {}'.format(
                        str(i+1).zfill(2), gini_train*100, new_var))
                    if show == 'both':
                        print('Step {} | 0:00:00.000000 | KS train = {:.2f}% | '
                        'Gini train = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), ks_train*100, gini_train*100, new_var))
                else:
                    if show == 'ks':
                        print('Step {} | 0:00:00.000000 | KS train = {:.2f}% | '
                        'KS test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), ks_train*100, ks_test*100, new_var))
                    if show == 'gini':
                        print('Step {} | 0:00:00.000000 | Gini train = {:.2f}% | '
                        'Gini test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), gini_train*100, gini_test*100, new_var))
                    if show == 'both':
                        print('Step {} | 0:00:00.000000 | KS train = {:.2f}% | '
                        'KS test = {:.2f}% | Gini train = {:.2f}% | Gini test = {:.2f}% '
                        '---> Feature selected: {}'.format(str(i+1).zfill(2), ks_train*100,
                        ks_test*100, gini_train*100, gini_test*100, new_var))

        else:

            if method == 'forward':

                if metric not in ('ks', 'gini'):
                    raise ValueError("El método 'forward' "
                    "solo se puede usar con las métricas 'ks' o 'gini'")

                start = datetime.datetime.now()

                contador = 0
                aux = pd.DataFrame(columns=['var', 'metric'])
                
                for var in var_list:

                    if var not in features:

                        features.append(var)
                        scorecard, features_length = compute_scorecard(
                        data, features, info, target_name=target_name, logistic_method=logistic_method, calibracion=calibracion)
                        data_final = apply_scorecard(data, scorecard, info, target_name)
                        metrica = compute_metrics(data_final, target_name, [metric])
                        aux.loc[contador] = [var, metrica]
                        features.pop()
                        contador += 1

                aux = aux.sort_values('metric', ascending=False)
                new_var = aux.iloc[0]['var']
                features.append(new_var)

                scorecard, features_length = compute_scorecard(
                data, features, info, target_name=target_name, logistic_method=logistic_method, calibracion=calibracion)
                train_final = apply_scorecard(data, scorecard, info, target_name)
                ks_train, gini_train = compute_metrics(train_final, target_name, ['ks', 'gini'])
                
                if metric == 'ks':
                    if ks_train <= old_ks+0.0020:
                        print('-' * N)
                        print('En el siguiente paso el KS no sube '
                        'más de un 0.20, detenemos el proceso')
                        features.pop()
                        break
                
                elif metric == 'gini':
                    if gini_train <= old_gini+0.0030:
                        print('-' * N)
                        print('En el siguiente paso el Gini no sube '
                        'más de un 0.30, detenemos el proceso')
                        features.pop()
                        break
                    
                old_ks, old_gini = ks_train, gini_train
                
                if not isinstance(muestra_test, type(None)):
                    test_final = apply_scorecard(muestra_test, scorecard, info, target_name)
                    ks_test, gini_test = compute_metrics(test_final, target_name, ['ks', 'gini'])

                if isinstance(muestra_test, type(None)):
                    end = datetime.datetime.now()
                    if show == 'ks':
                        print('Step {} | {} | KS train = {:.2f}% '
                        '---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, ks_train*100, new_var))
                    if show == 'gini':
                        print('Step {} | {} | Gini train = {:.2f}% '
                        '---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, gini_train*100, new_var))
                    if show == 'both':
                        print('Step {} | {} | KS train = {:.2f}% | '
                        'Gini train = {:.2f}% ---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, ks_train*100, gini_train*100, new_var))
                else:
                    end = datetime.datetime.now()
                    if show == 'ks':
                        print('Step {} | {} | KS train = {:.2f}% | '
                        'KS test = {:.2f}% ---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, ks_train*100, ks_test*100, new_var))
                    if show == 'gini':
                        print('Step {} | {} | Gini train = {:.2f}% | '
                        'Gini test = {:.2f}% ---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, gini_train*100, gini_test*100, new_var))
                    if show == 'both':
                        print('Step {} | {} | KS train = {:.2f}% | '
                        'KS test = {:.2f}% | Gini train = {:.2f}% | Gini test = {:.2f}% '
                        '---> Feature selected: {}'.format(str(i+1).zfill(2), end - start,
                        ks_train*100, ks_test*100, gini_train*100, gini_test*100, new_var))

            elif method == 'stepwise':

                if metric != 'pvalue':
                    raise ValueError("El método 'stepwise' "
                    "solo se puede usar con la métrica 'pvalue'")

                start = datetime.datetime.now()

                contador = 0
                aux = pd.DataFrame(columns=['var', 'pvalue', 'gini_auxiliar'])
                
                for var in var_list:

                    if var not in features:

                        features.append(var)
                        scorecard, features_length, pvalues = compute_scorecard(
                        data, features, info, target_name=target_name, pvalues=True, logistic_method=logistic_method, calibracion=calibracion)
                        pvalue = pvalues[-1]
                        if pvalue == 0:
                            gini_auxiliar = compute_metrics(apply_scorecard(
                            data, scorecard, info, target_name), target_name, ['gini'])
                        else: gini_auxiliar = 0
                        aux.loc[contador] = [var, pvalue, gini_auxiliar]
                        features.pop()
                        contador += 1

                aux = aux.sort_values(['pvalue', 'gini_auxiliar'], ascending=[True, False])
                best_pvalue = aux.iloc[0]['pvalue']

                if best_pvalue >= threshold:
                    print('-' * N)
                    print('Ya ninguna variable tiene un p-valor'
                    ' < {}, detenemos el proceso'.format(threshold))
                    break

                new_var = aux.iloc[0]['var']
                features.append(new_var)

                scorecard, features_length, pvalues = compute_scorecard(
                data, features, info, target_name=target_name, pvalues=True, logistic_method=logistic_method, calibracion=calibracion)
                new_pvalue = pvalues[-1]
                train_final = apply_scorecard(data, scorecard, info, target_name)
                ks_train, gini_train = compute_metrics(train_final, target_name, ['ks', 'gini'])
                
                if stop_ks_gini:
                    if ks_train <= old_ks+0.002 and gini_train <= old_gini+0.003:
                        print('-' * N)
                        print('En el siguiente paso ni el KS ni el GINI del train '
                        'suben más de un 0.20 o un 0.30 respectivamente, detenemos el proceso')
                        features.pop()
                        break
                old_ks, old_gini = ks_train, gini_train
                
                if not isinstance(muestra_test, type(None)):
                    test_final = apply_scorecard(muestra_test, scorecard, info, target_name)
                    ks_test, gini_test = compute_metrics(test_final, target_name, ['ks', 'gini'])

                if new_pvalue < 1e-100: pvalorcito = 0
                else: pvalorcito = new_pvalue

                if isinstance(muestra_test, type(None)):
                    end = datetime.datetime.now()
                    if show == 'ks':
                        print('Step {} | {} | pv = {:.2e} | '
                        'KS train = {:.2f}% ---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, pvalorcito, ks_train*100, new_var))
                    if show == 'gini':
                        print('Step {} | {} | pv = {:.2e} | '
                        'Gini train = {:.2f}% ---> Feature selected: {}'.format(
                        str(i+1).zfill(2), end - start, pvalorcito, gini_train*100, new_var))
                    if show == 'both':
                        print('Step {} | {} | pv = {:.2e} | '
                        'KS train = {:.2f}% | Gini train = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), end - start, pvalorcito,
                        ks_train*100, gini_train*100, new_var))

                else:
                    end = datetime.datetime.now()
                    if show == 'ks':
                        print('Step {} | {} | pv = {:.2e} | '
                        'KS train = {:.2f}% | KS test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), end - start, pvalorcito,
                        ks_train*100, ks_test*100, new_var))
                    if show == 'gini':
                        print('Step {} | {} | pv = {:.2e} | '
                        'Gini train = {:.2f}% | Gini test = {:.2f}% ---> Feature selected: {}'\
                        .format(str(i+1).zfill(2), end - start, pvalorcito,
                        gini_train*100, gini_test*100, new_var))
                    if show == 'both':
                        print('Step {} | {} | pv = {:.2e} | '
                        'KS train = {:.2f}% | KS test = {:.2f}% | Gini train = {:.2f}% | Gini test '
                        '= {:.2f}% ---> Feature selected: {}'.format(str(i+1).zfill(2),
                        end - start, pvalorcito, ks_train*100, ks_test*100,
                        gini_train*100, gini_test*100, new_var))

                dict_pvalues = dict(zip(features, pvalues[1:]))
                to_delete = {}
                for v in dict_pvalues:
                    if dict_pvalues[v] >= threshold:
                        to_delete[v] = dict_pvalues[v]
                if to_delete != {}:
                    for v in to_delete:
                        features.remove(v)
                        scorecard, features_length = compute_scorecard(
                        data, features, info, target_name=target_name, logistic_method=logistic_method, calibracion=calibracion)
                        train_final = apply_scorecard(data, scorecard, info, target_name)
                        ks_train, gini_train = compute_metrics(
                        train_final, target_name, ['ks', 'gini'])
                        old_ks, old_gini = ks_train, gini_train
                        if not isinstance(muestra_test, type(None)):
                            test_final = apply_scorecard(
                            muestra_test, scorecard, info, target_name)
                            ks_test, gini_test = compute_metrics(
                            test_final, target_name, ['ks', 'gini'])

                            if isinstance(muestra_test, type(None)):
                                end = datetime.datetime.now()
                                if show == 'ks':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | KS train = {:.2f}% ---> Feature deleted : {}'\
                                    .format(str(i+1).zfill(2), dict_pvalues[v], ks_train*100, v))
                                if show == 'gini':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | Gini train = {:.2f}% ---> Feature deleted : {}'\
                                    .format(str(i+1).zfill(2), dict_pvalues[v], gini_train*100, v))
                                if show == 'both':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | KS train = {:.2f}% | Gini train '
                                    '= {:.2f}% ---> Feature deleted : {}'.format(str(i+1).zfill(2),
                                    dict_pvalues[v], ks_train*100, gini_train*100, v))

                            else:
                                end = datetime.datetime.now()
                                if show == 'ks':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | KS train = {:.2f}% | KS test = {:.2f}% '
                                    '---> Feature deleted : {}'.format(str(i+1).zfill(2),
                                    dict_pvalues[v], ks_train*100, ks_test*100, v))
                                if show == 'gini':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | Gini train = {:.2f}% | Gini test = {:.2f}% '
                                    '---> Feature deleted : {}'.format(str(i+1).zfill(2),
                                    dict_pvalues[v], gini_train*100, gini_test*100, v))
                                if show == 'both':
                                    print('Step {} | 0:00:00.000000 | pv '
                                    '= {:.2e} | KS train = {:.2f}% | KS test = {:.2f}% | '
                                    'Gini train = {:.2f}% | Gini test = {:.2f}% ---> Feature '
                                    'deleted : {}'.format(str(i+1).zfill(2), dict_pvalues[v],
                                    ks_train*100, ks_test*100, gini_train*100, gini_test*100, v))

    print('-' * N)
    print('Selección terminada: {}'.format(features))
    print('-' * N)

    return features    


def table_ng(X, y, variable, objeto, bp):

    if not isinstance(bp, dict):
        vector = data_convert(X[variable].values, string_categories2(bp))[3]
        breakpoints_num = breakpoints_to_num(bp)
        groups_names = compute_group_names(objeto.dtype, bp)
        return compute_table(vector, y, breakpoints_num, groups_names)[0]

    else:
        vector = remapeo_missing(data_convert(X[variable].values, string_categories2(bp))[3], bp)
        breakpoints_num = breakpoints_to_num(bp['bp'])
        groups_names = compute_group_names(objeto.dtype, bp['bp'], bp['mg'])                                 
        return compute_table(vector, y, breakpoints_num, groups_names)[0]
        
        
def reagrupa_var(X, y, autogroupings, variable, user_bp={}, comparing_bp={},
color1='lightgreen', color2='lightgreen', color3='aquamarine', dims=(18, 5)):
    
    import matplotlib.pyplot as plt
    
    N = 100
    
    from IPython.display import display, HTML
    
    objeto = autogroupings[variable]
    
    if comparing_bp == {}:
        df1 = objeto.table
        original_bp = objeto.breakpoints
        color = color1

        
    else: 
        df1 = table_ng(X, y, variable, objeto, comparing_bp)
        original_bp = comparing_bp
        color = color3

    df1_html = df1.to_html()
    
    if user_bp == {}:
    
        html_str = f'''
        <div style="display: flex">
            <div style="margin-left: 00px; margin-right: 40px;">{df1_html}</div>
        </div>
        '''

        display(HTML(html_str))

    c1 = [str(i) for i in df1.index[:-1]]
    percent1 = df1['Percent'][:-1]
    badrate1 = df1['Bad rate'][:-1]

    fig = plt.figure(figsize=dims)

    plt.subplot(1, 2, 1)
    plt.bar(c1, percent1, color=color, width=0.5, label='Percent')
    plt.plot(c1, badrate1, color='red', linestyle='--', marker='o', label='Bad rate')
    plt.legend(loc='best')
    
    if comparing_bp == {}: plt.title('Dist. total pob. agrupación automática')
    else: plt.title('Dist. total pob. agrupación original')
    
    if user_bp != {}: 
        
        df2 = table_ng(X, y, variable, objeto, user_bp)
        df2_html = df2.to_html()
        
        html_str = f'''
        <div style="display: flex">
            <div style="margin-left: 00px; margin-right: 40px;">{df1_html}</div>
            <div style="margin-left: 00px; margin-right: 40px;">{df2_html}</div>
        </div>
        '''
        
        display(HTML(html_str))
        
        c2 = [str(i) for i in df2.index[:-1]]
        percent2 = df2['Percent'][:-1]
        badrate2 = df2['Bad rate'][:-1]
        
        plt.subplot(1, 2, 2)
        plt.bar(c2, percent2, color=color2, width=0.5, label='Percent')
        plt.plot(c2, badrate2, color='red', linestyle='--', marker='o', label='Bad rate')
        plt.title('Dist. total pob. agrupación propuesta')
        plt.legend(loc='best')
        
    plt.show()
    
    print('-'*N)
    if objeto.dtype != 'O':
        original_bp2 = original_bp
        try: original_bp2['bp'] = list(original_bp['bp'])
        except: pass
        if comparing_bp == {}:
            print('Agrupación automática: {}'.format(original_bp2))
        else:
            print('Agrupación original: {}'.format(original_bp2))
    else:
        L = original_bp
        if comparing_bp == {}: print('Agrupación automática: {}'.format(L))
        else: print('Agrupación original: {}'.format(L))
    print('-'*N)
    
    if user_bp != {}: 
        print('Agrupación final propuesta: {}'.format(user_bp))
        print('-'*N)

def interfaz_grafica(X, y, autogroupings, variables, comparing_bp={},
color1='lightgreen', color2='lightgreen', color3='aquamarine', dims=(18, 5)):  
    
    import matplotlib.pyplot as plt, ipywidgets as widgets, copy
    from IPython.display import display, HTML
    
    ##### VARIABLE #####
    
    def selector_variable(lista):
    
        global ig_lista_variables

        ig_lista_variables = widgets.Dropdown(
            options=lista,
            description='Variable:'
        )

        return ig_lista_variables

    def seleccionar_variable():

        global ig_variable_elegida
        ig_variable_elegida = ig_lista_variables.value
    
    ##### ACCIONES #####
        
    def agrupa_click(x):
    
        global ig_user_bp
        global ig_array
        global ig_visitas

        if ig_user_bp != {}: ig_array = ig_user_bp
        else: ig_user_bp = copy.deepcopy(autogroupings[ig_variable_elegida].breakpoints)
        
        grouped_ids = [int(cb.description) for cb in ig_checkbox_hbox.children if cb.value]

        if autogroupings[ig_variable_elegida].dtype not in ('O', 'bool'):
            if ig_user_bp['mg'] != None and ig_user_bp['mg'] > 0:
                grouped_ids = [i-1 for i in grouped_ids]
            array_num = ig_user_bp['bp']
            ini = int(grouped_ids[0])
            fin = int(grouped_ids[-1])
            ig_user_bp['bp'] = list(np.delete(array_num, slice(ini, fin)))
            ig_visitas[ig_variable_elegida] = ig_user_bp
        
        else:
            ig_array = ig_user_bp
            K = [i for i in ig_array if ig_array.index(i) in grouped_ids]
            for k in K: ig_array.remove(k)
            ig_array.insert(grouped_ids[0], [j for i in K for j in i])
            ig_user_bp = ig_array
            ig_visitas[ig_variable_elegida] = ig_user_bp

        with output:
            output.clear_output(wait=True)
            displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp)

    def separa_click(x):

        global ig_user_bp
        global ig_array
        global ig_visitas

        if ig_user_bp != {}: ig_array = ig_user_bp
        else: ig_user_bp = copy.deepcopy(autogroupings[ig_variable_elegida].breakpoints)
                
        array_num = ig_user_bp['bp']
        
        value = float(float(input_text1.value))
        if ig_user_bp['mg'] != None and ig_user_bp['mg'] > 0:
            if array_num.index(min([i for i in array_num if i > value])) < ig_user_bp['mg']:
                ig_user_bp['mg'] += 1
        array_num = np.append(array_num, [value])
        
        
        
        array_num.sort()
        ig_user_bp['bp'] = list(array_num)
        ig_visitas[ig_variable_elegida] = ig_user_bp

        with output:
            output.clear_output(wait=True)
            displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp)
            
    def missing_click(x):
        
        global ig_user_bp
        global ig_array
        global ig_visitas

        if ig_user_bp != {}: ig_array = ig_user_bp
        else: ig_array = copy.deepcopy(autogroupings[ig_variable_elegida].breakpoints)
        
        array_num = list(ig_array['bp'])
        missing_group = int(input_text2.value)
        
        if missing_group != 0:
            try: array_num.remove(-12345670.0)
            except: pass
            
        if missing_group == 0:
            if -12345670.0 not in array_num: array_num.insert(0, -12345670.0)
        
        ig_user_bp['mg'] = missing_group
        ig_user_bp['bp'] = array_num
        ig_visitas[ig_variable_elegida] = ig_user_bp
        
        with output:
            output.clear_output(wait=True)
            displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp)        
            
    def desplaza_click(x):
        
        global ig_user_bp
        global ig_array
        global ig_visitas
        
        if ig_user_bp != {}: ig_array = ig_user_bp
        else: ig_array = copy.deepcopy(autogroupings[ig_variable_elegida].breakpoints)
        
        valor, grupo = input_text3.value, int(input_text4.value)
        K = copy.deepcopy(ig_array)
        if valor not in [j for i in K for j in i]: print('Ese valor no existe!')
        else:
            for i in ig_array:
                if valor in i:
                    if ig_array.index(i) == grupo: print('Ese valor ya está en ese grupo!')
                    else:
                        if len(i) > 1:
                            K[K.index(i)].remove(valor)
                            if grupo >= len(K): K.append([valor])
                            else: K[grupo].append(valor)
                        else:
                            if grupo >= len(K): K.append([valor])
                            else: K[grupo].append(valor)
                            K.remove(i)
                        
        ig_array = K
        ig_user_bp = ig_array
        ig_visitas[ig_variable_elegida] = ig_user_bp

        with output:
            output.clear_output(wait=True)
            displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp)
            
    def ir_click(x):
        
        global ig_contador
        global ig_user_bp
        global ig_array
        global ig_visitas
        
        with output:
            ig_array = []
            seleccionar_variable()
            if ig_contador == 0 and ig_variable_elegida in comparing_bp:
                ig_user_bp = copy.deepcopy(comparing_bp[ig_variable_elegida])
                ig_contador += 1
            else: ig_user_bp = ig_visitas[ig_variable_elegida]
            output.clear_output(wait=True)
            displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp)
            
    ##### DISPLAYER ######
            
    def displayer(X, y, autogroupings, ig_variable_elegida, ig_user_bp):
    
        layout_100 = widgets.Layout(width='100%')
        layout_95 = widgets.Layout(width='95%')

        global ig_checkbox_hbox
        global agrupaciones_interfaz

        formato = f'''
        <h1 style="
        text-align:center;
        font-size: 12pt;
        background-color: lightgray;
        margin: 0;
        padding: 10px
        ">{ig_variable_elegida}</h1>
        '''

        objeto = autogroupings[ig_variable_elegida]
            
        if ig_variable_elegida not in comparing_bp:
            df1 = objeto.table
            color = color1
        else:  
            comparing_bp_var = copy.deepcopy(comparing_bp[ig_variable_elegida])
            df1 = me.table_ng(X, y, ig_variable_elegida, objeto, comparing_bp_var)
            color = color3
            
        df1_html = df1.to_html()
            
        if ig_user_bp == {}:
            ig_checkbox_hbox = widgets.HBox([widgets.Checkbox(value=False, description=str(x)) for x in df1.index[:-1]])

        title_widget = widgets.HTML(value=formato, layout=layout_95)

        c1 = [str(i) for i in df1.index[:-1]]
        percent1 = df1['Percent'][:-1]
        badrate1 = df1['Bad rate'][:-1]

        fig = plt.figure(figsize=dims)
        
        plt.subplot(1, 2, 1)
        plt.bar(c1, percent1, color=color, width=0.5, label='Percent')
        plt.plot(c1, badrate1, color='red', linestyle='--', marker='o', label='Bad rate')
        if ig_variable_elegida in comparing_bp: plt.title('Dist. total pob. agrupación automática')
        else: plt.title('Dist. total pob. agrupación original')
        plt.legend(loc='best')
        
        if ig_user_bp == {}:
            
            html_str = f'''
            <div style="display: flex">
                <div style="margin-left: 00px; margin-right: 40px;">{df1_html}</div>
            </div>
            '''
            
        if ig_user_bp != {}:

            df2 = me.table_ng(X, y, ig_variable_elegida, objeto, ig_user_bp)
            if objeto.dtype not in ('O', 'bool') and  ig_user_bp['mg'] != None and ig_user_bp['mg'] > 0:
                df2.index = [i+1 for i in df2.index[:-1]] + [df2.index[-1]]
            df2_html = df2.to_html()
            
            ig_checkbox_hbox = widgets.HBox([widgets.Checkbox(value=False, description=str(x)) for x in df2.index[:-1]])

            html_str = f'''
            <div style="display: flex">
                <div style="margin-left: 00px; margin-right: 40px;">{df1_html}</div>
                <div style="margin-left: 00px; margin-right: 40px;">{df2_html}</div>
            </div>
            '''

            c2 = [str(i) for i in df2.index[:-1]]
            percent2 = df2['Percent'][:-1]
            badrate2 = df2['Bad rate'][:-1]

            plt.subplot(1, 2, 2)
            plt.bar(c2, percent2, color=color2, width=0.5, label='Percent')
            plt.plot(c2, badrate2, color='red', linestyle='--', marker='o', label='Bad rate')
            plt.title('Dist. total pob. agrupación propuesta')
            plt.legend(loc='best')
            
        if objeto.dtype not in ('O', 'bool'):
            display(widgets.VBox([title_widget]))
            display(widgets.HBox([widgets.HBox([agrupar_button]), widgets.HBox([ig_checkbox_hbox])], layout=widgets.Layout(width='70%')))
            display(widgets.HBox([widgets.HBox([separar_button]), widgets.HBox([input_text1])], layout=layout_100))
            if objeto.missing:
                display(widgets.HBox([widgets.HBox([missing_button]), widgets.HBox([input_text2])], layout=layout_100))

        else:            
            display(widgets.VBox([title_widget])) 
            display(widgets.HBox([widgets.HBox([agrupar_str_button]), widgets.HBox([ig_checkbox_hbox])], layout=layout_100))
            display(widgets.HBox([widgets.HBox([desplazar_button]), widgets.HBox([input_text3]), widgets.HBox([input_text4])], layout=layout_100))        

            
        display(HTML(html_str))

        plt.show()
        
        agrupaciones_interfaz = {i:ig_visitas[i] for i in ig_visitas if ig_visitas[i] != {}}
        
    ##### BOTONES ######
    
    ir_button = widgets.Button(description='Ir')
    ir_button.on_click(ir_click)

    agrupar_button = widgets.Button(description='Agrupar')
    agrupar_button.on_click(agrupa_click)

    separar_button = widgets.Button(description='Separar')
    separar_button.on_click(separa_click)

    missing_button = widgets.Button(description='Missing')
    missing_button.on_click(missing_click)
    
    agrupar_str_button = widgets.Button(description='Agrupar')
    agrupar_str_button.on_click(agrupa_click)
    
    desplazar_button = widgets.Button(description='Desplazar')
    desplazar_button.on_click(desplaza_click)
    
    ##### TEXTOS #####
    
    input_text1 = widgets.Text(description='Valor')
    input_text2 = widgets.Text(description='Grupo')
    input_text3 = widgets.Text(description='Valor')
    input_text4 = widgets.Text(description='Grupo')
    
    ##### EJECUCIÓN #####
    
    global ig_visitas
    ig_visitas = {}
    for var in variables: ig_visitas[var] = {}
    
    output = widgets.Output()
    display(widgets.HBox([selector_variable(variables), widgets.HBox([ir_button])]))
    display(output)
    
    global ig_contador
    ig_contador = 0
    
    global agrupaciones_interfaz
    agrupaciones_interfaz = {i:ig_visitas[i] for i in ig_visitas if ig_visitas[i] != {}}

def pretty_scorecard(modelo, color1='blue', color2='#FFFFFF'):
    
    if color1 == 'green': color1 = '#CCFFCC'
    if color1 == 'light_blue': color1 = '#CCFFFF'
    if color1 == 'blue': color1 = '#CCECFF'
    if color1 == 'pink': color1 = '#FFCCFF'
    if color1 == 'red': color1 = '#FFCCCC'
    if color1 == 'yellow': color1 = '#FFFFCC'
    if color1 == 'purple': color1 = '#CCCCFE'
    if color1 == 'orange': color1 = '#FFCC99'

    contador1, contador2, indices1, indices2 =  0, 0, [], []
    for i in modelo.features_length:
        for j in range(i):
            if contador1 % 2 == 0: indices1.append(contador2+j)
            else: indices2.append(contador2+j)
        contador1, contador2 = contador1+1, contador2+i

    def row_style(row):
        if row.name in indices1: return pd.Series('background-color: {}'.format(color1), row.index)
        else: return pd.Series('background-color: {}'.format(color2), row.index)

    try:display(modelo.scorecard.style.apply(row_style, axis=1))
    except: display(modelo.scorecard)


def parceling(df_in, breakpoints=[], tramos=15, id_columns=['id'],
score_name='scorecardpoints_acep', target_name='target', randomly=True):

    df = df_in.copy()

    if randomly: np.random.seed(123)

    if breakpoints == []:

        tabla = proc_freq(df, score_name)

        inf = min(tabla.index)
        sup = max(tabla.index)
        salto = (sup - inf) / tramos
        breakpoints = [round(inf+i*salto, 2) for i in range(tramos)]

    df['parcel'] = np.digitize(df[score_name], breakpoints)
    a = proc_freq(df, 'parcel', target_name)
    a.columns.name = None
    a = a.reset_index(drop=True)
    a.index.name = 'parcel'
    b = proc_freq(df[df[target_name].isin([0, 1])], 'parcel',
    target_name, option='pct_row')
    b.columns.name = None
    b = b.reset_index(drop=True)
    b.index.name = 'parcel'
    b = b.rename(columns={0: '0_pct', 1: '1_pct'})
    c = a.merge(b, on='parcel', how='left')
    contador = 0
    molde = pd.DataFrame()
    for i in c.index:
        Xaux = df[(df['parcel'] == i+1) & (df['decision'].isin(['denegado', 'rechazado']))].copy()
        mascaritaaa = np.array([True]*round(len(Xaux)*c.loc[i]['1_pct'])
        +[False]*(len(Xaux)-round(len(Xaux)*c.loc[i]['1_pct'])))
        if randomly: np.random.shuffle(mascaritaaa)
        else: Xaux = Xaux.sort_values(score_name)
        Xaux['target_inf'] = np.where(mascaritaaa, 1, 0)
        contador += len(Xaux)
        molde = pd.concat([molde, Xaux])
    df2 = df.merge(molde[id_columns + ['target_inf']], how='left', on=id_columns)
    df2['target_def'] = np.where(df2['target_inf'].isna(), df2[target_name], df2['target_inf'])
    
    return df2, c


def cell_style(cell, name='Calibri', size=11, bold=False, italic=False, underline='none',
font_color='FF000000', background_color='', all_borders=False, hor_alignment='general',
ver_alignment='bottom', wrap_text=False, left_border=None, right_border=None, top_border=None,
bottom_border=None, left_border_color='FF000000', right_border_color='FF000000',
top_border_color='FF000000', bottom_border_color='FF000000'):
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if background_color != '':
         fill_type = 'solid'
    else:
        background_color = 'FF000000'
        fill_type = None

    if all_borders == True:
        left_border, right_border, top_border, bottom_border = 'thin', 'thin', 'thin', 'thin'

    cell.font = Font(name=name, size=size, bold=bold,
    italic=italic, underline=underline, color=font_color)
    cell.fill = PatternFill(fill_type=fill_type, fgColor=background_color)
    cell.alignment = Alignment(horizontal=hor_alignment,
    vertical=ver_alignment, wrap_text=wrap_text)
    cell.border = Border(left=Side(border_style=left_border, color=left_border_color),
    right=Side(border_style=right_border, color=right_border_color),
    top=Side(border_style=top_border, color=top_border_color),
    bottom=Side(border_style=bottom_border, color=bottom_border_color))


def predict_pyspark(data, features, pyspark_formula,
target_name='', keep_columns=[], binary_treshold=0.0, score_name='scorecardpoints'):
            
    import pyspark.sql.functions as sf
    from pyspark.sql.types import IntegerType
    
    if target_name != '': X1 = data.select(keep_columns + features + [target_name])
    else: X1 = data.select(keep_columns + features)
            
    X1 = X1.withColumn(score_name, sf.lit(0).cast(IntegerType()))

    for i in range(len(pyspark_formula)):
        X1 = X1.withColumn('scr_{}'.format(features[i]), 
        sf.expr(pyspark_formula[i])).withColumn(score_name,
        sf.col(score_name) + sf.col('scr_{}'.format(features[i])))

    if binary_treshold != 0.0:
        X1 = X1.withColumn('prediction',
        sf.when(sf.col(score_name) >= binary_treshold, 0).otherwise(1))

    columnas = list(X1.columns).copy()
    columnas.remove(score_name)
    if binary_treshold != 0.0: columnas.remove('prediction')
    columnas += [score_name]
    if binary_treshold != 0.0: columnas += ['prediction']
    X1 = X1.select(columnas)

    return X1


def metrics_pyspark(data, target_name, metrics, print_log=False, score_name='scorecardpoints'):
    
    import pyspark.sql.functions as sf
    from pyspark.sql.types import DoubleType
    from pyspark.ml.evaluation import BinaryClassificationEvaluator

    if metrics not in (['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']):
        raise ValueError("Valor erroneo para 'metrics'. Los valores "
        "váidos son: ['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']")

    if 'ks' in metrics:
        ks = compute_pyspark_ks(data.withColumn(score_name,
        sf.col(score_name).cast(DoubleType())), target_name, score_name)

    if 'gini' in metrics:
        evaluator = BinaryClassificationEvaluator(rawPredictionCol=score_name,
        labelCol=target_name, metricName='areaUnderROC')
        auroc = evaluator.evaluate(data.withColumn(
        score_name, sf.col(score_name).cast(DoubleType())))
        gini = 1 - 2 * auroc

    if metrics == ['ks']:
        if print_log:
            print('El modelo tiene un {:.2f}% de KS en esta muestra'.format(round(ks*100, 2)))
        return ks

    if metrics == ['gini']:
        if print_log:
            print('El modelo tiene un {:.2f}% de Gini en esta muestra'.format(round(gini*100, 2), ))
        return gini

    if metrics in (['ks', 'gini'], ['gini', 'ks']):
        if print_log:
            print('El  modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini '
            'en esta muestra'.format(round(ks*100, 2), round(gini*100, 2)))
        return ks, gini
    
    
def compute_pyspark_ks(df, target, score):
    
    import pyspark.sql.functions as sf
    from pyspark.ml.feature import Bucketizer
    
    df = df.withColumn(score, sf.round(sf.col(score), 3))
    minimo = df.agg({score: 'min'}).collect()[0][0]
    maximo = df.agg({score: 'max'}).collect()[0][0]
    bins = np.arange(minimo, maximo + 0.001, np.round((maximo - minimo) / 1000, 3))
    bins[0] = -float('inf')
    bins[len(bins) -1 ] = float('inf')

    bucketizer = Bucketizer(splits=list(bins), inputCol=score, outputCol='buckets')
    bucketed = bucketizer.setHandleInvalid('keep').transform(df)
    pre_pivot = bucketed.groupby('buckets', target).count().toPandas()
    pivot_table = pre_pivot.pivot_table(values='count', columns=target, index='buckets').fillna(0)
    pivot_table['pct_ceros'] = pivot_table.iloc[:, 0] / np.sum(pivot_table.iloc[:, 0].values)
    pivot_table['pct_unos'] = pivot_table.iloc[:, 1] / np.sum(pivot_table.iloc[:, 1].values)
    pivot_table['pct_ceros_cum'] = pivot_table['pct_ceros'].cumsum()
    pivot_table['pct_unos_cum'] = pivot_table['pct_unos'].cumsum()
    pivot_table['KS'] = (pivot_table['pct_ceros_cum'] - pivot_table['pct_unos_cum']).abs()
    KS = pivot_table['KS'].max()

    return KS
    

def save_model(modelo, name):

    import _pickle

    pickle_file = open(name, 'wb')
    _pickle.dump(modelo, pickle_file)
    pickle_file.close()
    
    
def save_light_model(modelo, name):

    import _pickle

    light_model = {
        'features': modelo.features,
        'pyspark_formula': modelo.pyspark_formula,
        }

    pickle_file = open(name, 'wb')
    _pickle.dump(light_model, pickle_file)
    pickle_file.close()
    
    
def load_model(name):

    import _pickle

    pickle_file = open(name, 'rb')
    obj = _pickle.load(pickle_file)
    pickle_file.close()
    
    return obj


def genera_punt_par(df, features):
    
    import pyspark.sql.functions as sf
    
    for feature in features:
        condition = 'sf'
        for rc in [i for i in dfOut.columns if 'reasoncode_' in i]:
            condition += ".when(sf.col('{}').contains('{}'), sf.col('{}'))".format(rc, feature, rc)
        df = df.withColumn('grp_{}'.format(feature), eval(condition))
    
    return df


def proc_freq(data, row, col='', weight='', decimals=None, cumulative=False,
sort_col='', sort_dir='', option='', values=[], output=None):

    '''
    Generates the frequency table of a variable in a DataFrame. If two variables are passed,
    inside the 'row' and 'col' parameters, then it computes their crosstab.
    :param data: DataFrame. Table to use. Supports both pandas and spark Dataframe.
    :param row: str. Column to compute its frequency table.
    :param col: str. Column to compute its crosstab combined with 'row'.
    :param weight: str. Column with the frequencies of the distinct 'row' values.
    :param decimals: int. Decimal precision. Not rounded by default.
    :param sort_col: str. Column to sort by. It's sorted ascending on index by default.
    :param sort_dir: str. Direction to sort by. Use 'desc' for descending order.
    :param cumulative: bool. If True then returns cumulative frequency and percentage.
    :param option: str. By default, the crosstabs are computed with frequencies.
    Use 'pct_row' or 'pct_col' to obtain the desire percentages in crosstabs.
    :param values: list. In a frequency table as a pandas.DataFrame,
    it shows all the values of the list filling the ones that do not appear with zeros.
    :param output: SparkSession. By default the function returns a pandas.DataFrame.
    Input your spark session if a spark.DataFrame is wanted.
    :return:
    '''

    if type(data) == type(pd.DataFrame([])): # pandas.DataFrame

        if col == '': # Frequency table

            if weight == '': freq = data.groupby(row, dropna=False).size().to_frame()
            else: freq = data.groupby(row, dropna=False).agg({weight: 'sum'})
            freq.columns = ['frequency']

            if decimals == None: freq['percent'] = freq['frequency'] / freq['frequency'].sum()
            else: freq['percent'] = (freq['frequency'] / freq['frequency'].sum()).round(decimals)

            if sort_col == '' or sort_col == row:
                if sort_dir == 'desc': freq = freq.sort_index(ascending=False)
            else:
                if sort_dir == 'desc': freq = freq.sort_values(sort_col, ascending=False)
                else: freq = freq.sort_values(sort_col)

            if cumulative == True:
                freq['cumulative_frequency'] = freq['frequency'].cumsum()
                if decimals == None:
                    freq['cumulative_percent'] = \
                    (freq['frequency'] / freq['frequency'].sum()).cumsum()
                else:
                    freq['cumulative_percent'] = \
                    ((freq['frequency'] / freq['frequency'].sum()).cumsum()).round(decimals)

            if output != None:
                freq = freq.reset_index()
                freq = output.createDataFrame(freq)

        else: # Crosstab

            dataaa = data.copy()
            dataaa[row], dataaa[col] = dataaa[row].fillna(np.e), dataaa[col].fillna(np.e)
            freq = pd.pivot_table(dataaa, index=[row], columns=[col], aggfunc='size',
            fill_value=0).rename(columns={np.e: np.nan}, index={np.e: np.nan})

            if option == 'pct_col':
                for column in freq.columns:
                    if decimals == None: freq[column] = freq[column] / freq[column].sum()
                    else: freq[column] = (freq[column] / freq[column].sum()).round(decimals)

            if option == 'pct_row':
                suma = freq.sum(axis=1)
                for column in freq.columns:
                    if decimals == None: freq[column] = freq[column] / suma
                    else: freq[column] = (freq[column] / suma).round(decimals)

            if sort_col == '' or sort_col == row:
                if sort_dir == 'desc': freq = freq.sort_index(ascending=False)
            else:
                if sort_dir == 'desc': freq = freq.sort_values(sort_col, ascending=False)
                else: freq = freq.sort_values(sort_col)

            if output != None:
                freq.columns.names = [None]
                freq = freq.reset_index()
                freq = output.createDataFrame(freq)
                freq = freq.withColumnRenamed(row, row + '_' + col)

    else: # pyspark.DataFrame
        
        import pyspark.sql.functions as sf
        from pyspark.sql.types import IntegerType, FloatType
        from pyspark.sql.window import Window

        if col == '': # Frequency table

            freq = data.groupBy(row).count().withColumnRenamed('count', 'frequency')
            freq = freq.sort(row)

            if output != None:

                suma = freq.agg(sf.sum('frequency')).collect()[0][0]
                if decimals == None:
                    freq = freq.withColumn('percent',
                    sf.col('frequency') / sf.lit(suma))
                else:
                    freq = freq.withColumn('percent',
                    sf.format_number(sf.col('frequency') / sf.lit(suma), decimals))

                if sort_col == '':
                    if sort_dir == 'desc': freq = freq.sort(row, ascending=False)
                else:
                    if sort_dir == 'desc': freq = freq.sort(sort_col, ascending=False)
                    else: freq = freq.sort(sort_col)

                if cumulative == True:
                    freq = freq.withColumn('cumulative_frequency',
                    sf.sum('frequency').over(Window.rowsBetween(Window.unboundedPreceding, 0)))
                    if decimals == None: freq = freq.withColumn('cumulative_percent',
                    sf.sum(sf.col('frequency') / sf.lit(suma))\
                    .over(Window.rowsBetween(Window.unboundedPreceding, 0)))
                    else: freq = freq.withColumn('cumulative_percent',
                    sf.format_number(sf.sum(sf.col('frequency') / sf.lit(suma))\
                    .over(Window.rowsBetween(Window.unboundedPreceding, 0)), decimals))

            else:

                freq = freq.toPandas().set_index(row)

                if decimals == None: freq['percent'] = freq['frequency'] / freq['frequency'].sum()
                else: 
                    freq['percent'] = (freq['frequency'] / freq['frequency'].sum()).round(decimals)

                if sort_col == '' or sort_col == row:
                    if sort_dir == 'desc': freq = freq.sort_index(ascending=False)
                else:
                    if sort_dir == 'desc': freq = freq.sort_values(sort_col, ascending=False)
                    else: freq = freq.sort_values(sort_col)

                if cumulative == True:
                    freq['cumulative_frequency'] = freq['frequency'].cumsum()
                    if decimals == None:
                        freq['cumulative_percent'] = \
                        (freq['frequency'] / freq['frequency'].sum()).cumsum()
                    else:
                        freq['cumulative_percent'] = \
                        ((freq['frequency'] / freq['frequency'].sum()).cumsum()).round(decimals)

        else: # Crosstab

            freq = data.crosstab(row, col)

            if data.select(row).dtypes[0][1] in ('smallint', 'int', 'bigint'):
                freq = freq.withColumn(row + '' + col, sf.col(row + '' + col).cast(IntegerType()))
            elif data.select(row).dtypes[0][1] == 'double':
                freq = freq.withColumn(row + '' + col, sf.col(row + '' + col).cast(FloatType()))

            if data.select(col).dtypes[0][1] in ('smallint', 'int', 'bigint'):
                L1, L2 = [], []
                for i in freq.columns[1:]:
                    try: L1.append(int(i))
                    except: L2.append(i)
                L1.sort()
                L3 = L2 + [str(i) for i in L1]
                freq = freq.select([freq.columns[0]] + L3)
            elif data.select(col).dtypes[0][1] == 'double':
                L1, L2 = [], []
                for i in freq.columns[1:]:
                    try: L1.append(float(i))
                    except: L2.append(i)
                L1.sort()
                L3 = L2 + [str(i) for i in L1]
                freq = freq.select([freq.columns[0]] + L3)

            freq = freq.sort(row + '_' + col)

            if output != None:

                if option == 'pct_col':
                    for column in list(freq.columns[1:]):
                        if decimals == None: freq = freq.withColumn(
                        column, sf.col(column) / sf.sum(column).over(Window.partitionBy()))
                        else: freq = freq.withColumn(
                        column, sf.format_number(sf.col(column) / sf.sum(column)\
                        .over(Window.partitionBy()), decimals))

                if option == 'pct_row':
                    for column in list(freq.columns[1:]):
                        if decimals == None:
                            freq = freq.withColumn(column,
                            sf.col(column) / sum([sf.col(c) for c in freq.columns[1:]]))
                        else:
                            freq = freq.withColumn(column,
                            sf.format_number(sf.col(column) / sum([sf.col(c)
                            for c in freq.columns[1:]]), decimals))

                if sort_col == '':
                    if sort_dir == 'desc': freq = freq.sort(row + '_' + col, ascending=False)
                else:
                    if sort_dir == 'desc': freq = freq.sort(sort_col, ascending=False)
                    else: freq = freq.sort(sort_col)

            else:

                freq = freq.toPandas()
                freq = freq.rename(columns={row + '_' + col: row})
                freq = freq.set_index(row)
                freq.columns.name = col

                if option == 'pct_col':
                    for column in freq.columns:
                        if decimals == None: freq[column] = freq[column] / freq[column].sum()
                        else: freq[column] = (freq[column] / freq[column].sum()).round(decimals)

                if option == 'pct_row':
                    denominador = freq.sum(axis=1)
                    for column in freq.columns:
                        if decimals == None: freq[column] = freq[column] / denominador
                        else: freq[column] = (freq[column] / denominador).round(decimals)

                if sort_col == '' or sort_col == row:
                    if sort_dir == 'desc': freq = freq.sort_index(ascending=False)
                else:
                    if sort_dir == 'desc': freq = freq.sort_values(sort_col, ascending=False)
                    else: freq = freq.sort_values(sort_col)

    if type(freq) == type(pd.DataFrame([])) and len(values) > 0:

        for value in values:
            if value not in freq.index:
                freq.loc[value] = [0]*len(freq.columns)

        if sort_col == '' or sort_col == row:
            if sort_dir == 'desc': freq = freq.sort_index(ascending=False)
            else: freq = freq.sort_index() # Necesita reordenar sí o sí
        else:
            if sort_dir == 'desc': freq = freq.sort_values(sort_col, ascending=False)
            else: freq = freq.sort_values(sort_col)

    return freq

class Scorecard:


    def __init__(
        
        self,
        
        test_seed=123,
        test_size=0.3,
        stratify=True,
        stratify_var='',
        flag_train_test=[],

        autogrp_all=False,
        autogrp_max_groups=5,
        autogrp_min_pct=0.05,
        autogrp_dict_max_groups={},
        autogrp_dict_min_pct={},
        autogrp_dict_manual_types={},

        features=[],
        excluded_vars=[],
        included_vars=[],

        iv_threshold=0.015,
        selection_show = 'gini',
        selection_method='stepwise',
        selection_metric='pvalue',
        selection_threshold=0.01,
        selection_max_iters=14,
        selection_stop_ks_gini=True,
        selection_check_overfitting=True,
        
        calibracion='default',
        user_breakpoints={},
        logistic_method = 'lbfgs',

        id_columns=[],
        save_tables='features',
        save_autogroupings='features',
        warnings='disabled',
        
        verbose=True
        
    ):

        self.test_seed = test_seed
        self.test_size = test_size
        self.stratify = stratify
        self.stratify_var = stratify_var
        self.flag_train_test = flag_train_test
        
        self.autogrp_all = autogrp_all
        self.autogrp_max_groups = autogrp_max_groups
        self.autogrp_min_pct = autogrp_min_pct
        self.autogrp_dict_max_groups = autogrp_dict_max_groups
        self.autogrp_dict_min_pct = autogrp_dict_min_pct
        self.autogrp_dict_manual_types = autogrp_dict_manual_types
        
        self.features = features
        self.excluded_vars = excluded_vars
        self.included_vars = included_vars
        
        self.iv_threshold = iv_threshold
        self.selection_show = selection_show
        self.selection_method = selection_method
        self.selection_metric = selection_metric
        self.selection_threshold = selection_threshold
        self.selection_max_iters = selection_max_iters
        self.selection_stop_ks_gini = selection_stop_ks_gini
        self.selection_check_overfitting = selection_check_overfitting
        
        self.calibracion = calibracion
        self.user_breakpoints = user_breakpoints
        self.logistic_method = logistic_method
        
        self.id_columns = id_columns
        self.save_tables = save_tables
        self.save_autogroupings = save_autogroupings
        self.warnings = warnings
        
        self.verbose = verbose
        
        if self.warnings == 'disabled':
            
            import warnings
            warnings.filterwarnings('ignore')
        

    def fit(self, X, y):
        
        if y.dtype == 'O':
            raise Exception("CUIDADO! La columna con el target (y) debe ir en tipo entero y \
solo con 0's y 1's. Si está en tipo 'object' prueba a cambiarla con '.astype(int)'")

        ##### SPLIT
        
        X_train, X_test, y_train, y_test, index_train, index_test = split(
        X, y, self.test_seed, self.test_size, self.stratify,
        self.stratify_var, self.flag_train_test, self.verbose)

        ##### AUTOGROUPING
                
        if self.autogrp_all: variables = X.columns
        else: 
            if self.features != []: variables = self.features
            else: variables = list(set(list(X.columns)) - set(self.excluded_vars) - set(self.id_columns))
                
        autogroupings, variables_no_agrupadas_error = autogrouping(
        X_train, y_train, variables, self.autogrp_max_groups,
        self.autogrp_min_pct, self.autogrp_dict_max_groups,
        self.autogrp_dict_min_pct, self.autogrp_dict_manual_types, self.verbose)
        
        ##### DATA TRANSFORM
        
        if self.features != []: variables_semidef = self.features
        else: variables_semidef = [i for i in autogroupings]

        self.final_breakpoints = compute_final_breakpoints(
        variables_semidef, autogroupings, self.user_breakpoints)
        info = compute_info(X_train, variables_semidef, self.final_breakpoints)

        df_train = adapt_data(X_train, y_train, variables_semidef, self.final_breakpoints)
        df_test = adapt_data(X_test, y_test, variables_semidef, self.final_breakpoints)
        
        ##### UPDATE IV
        
        for i in self.user_breakpoints:          
            x = df_train[i].values
            breakpoints_num = info[i]['breakpoints_num']
            group_names = info[i]['group_names']
            autogroupings[i].iv = compute_table(x, y_train, breakpoints_num, group_names, False)[1]
            
        ##### IV THRESHOLD
        
        tabla_ivs, contador = pd.DataFrame(columns=['variable', 'iv']), 0
        for variable in autogroupings:
            tabla_ivs.loc[contador] = variable, autogroupings[variable].iv
            contador += 1
        self.tabla_ivs = tabla_ivs.sort_values('iv', ascending=False).reset_index(drop=True)
        variables_def = tabla_ivs[tabla_ivs['iv'] >= self.iv_threshold]['variable']
        
        ##### FEATURE SELECTION

        features = features_selection(
        df_train, self.features, variables_def, info, 'target_4815162342',
        method=self.selection_method, metric=self.selection_metric,
        threshold=self.selection_threshold, stop_ks_gini=self.selection_stop_ks_gini,
        max_iters=self.selection_max_iters, included_vars=self.included_vars,
        muestra_test=df_test, show=self.selection_show, 
        check_overfitting=self.selection_check_overfitting,
        logistic_method=self.logistic_method, calibracion=self.calibracion)

        df_train = df_train[features + ['target_4815162342']]
        df_test = df_test[features + ['target_4815162342']]

        scorecard, features_length, pvalues, coefs = compute_scorecard(
        df_train, features, info, pvalues=True, 
        ret_coefs=True, logistic_method=self.logistic_method, calibracion=self.calibracion)

        df_train_final = apply_scorecard(df_train, scorecard, info, 'target_4815162342')
        ks_train, gini_train = compute_metrics(df_train_final, 'target_4815162342', ['gini', 'ks'])

        print('El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de entrenamiento'.format(round(ks_train*100, 2), round(gini_train*100, 2)))
        print('-' * N)

        df_test_final = apply_scorecard(df_test, scorecard, info, 'target_4815162342')
        ks_test, gini_test = compute_metrics(df_test_final, 'target_4815162342', ['gini', 'ks']) 

        print('El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de validación'.format(round(ks_test*100, 2), round(gini_test*100, 2)))
        print('-' * N)
        
        self.index_train = index_train
        self.index_test = index_test
        
        self.variables_no_agrupadas_error = variables_no_agrupadas_error

        self.features = features
        self.scorecard = scorecard
        self.features_length = features_length
        self.pvalues = dict(zip(features, list(pvalues[1:])))
        self.coefs = coefs

        self.ks_train = ks_train
        self.gini_train = gini_train
        self.ks_test = ks_test
        self.gini_test = gini_test

        if self.save_tables == 'all': 
            self.X_train = X_train
            self.y_train = y_train
            self.X_test = X_test
            self.y_test = y_test
        elif self.save_tables == 'features':
            self.X_train = X_train[self.id_columns + features]
            self.y_train = y_train
            self.X_test = X_test[self.id_columns + features]
            self.y_test = y_test
            
        for objeto in autogroupings: del autogroupings[objeto].x_final

        if self.save_autogroupings == 'all': 
            self.autogroupings = autogroupings
        elif self.save_autogroupings == 'features':
            self.autogroupings = dict((k, autogroupings[k]) for k in features if k in autogroupings)
            
        try: self.create_pyspark_formula()
        except: self.pypsark_formula = ['Por algún motivo no se ha podido calcular la fórmula SQL']

        return self

    def create_excel(self, ruta, color='blue'):
        
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill
        
        scorecard = self.scorecard.copy()
        scorecard = scorecard.drop('Raw score', axis=1)
        
        abc = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
        'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
        'AD', 'AE', 'AF', 'AG','AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP',
        'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        rows = dataframe_to_rows(scorecard, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                try: ws.cell(row=r_idx, column=c_idx, value=value)
                except: ws.cell(row=r_idx, column=c_idx, value=str(value))

        ws.insert_cols(2)
        ws.insert_cols(4)
        ws.insert_cols(12)

        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:D1')

        altura = len(scorecard)

        for letra in ['F', 'I']:
            for row in ws['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.00%'

        for letra in ['J', 'K', 'L']:
            for row in ws['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.0000'

        for i in range(2, altura+2):
            ws.merge_cells('C{}:D{}'.format(i, i))

        for row in ws['A1:M1']:
            for cell in row:
                cell_style(cell, bold=True, hor_alignment='center', ver_alignment='center', 
                all_borders=True, font_color='ffffff', background_color='ff0000')

        for row in ws['A2:M{}'.format(altura+1)]:
            for cell in row:
                cell_style(cell, hor_alignment='center', 
                ver_alignment='center', all_borders=True, wrap_text=True)

        ws['K1'].value = 'IV aux'
        ws['L1'].value = 'IV'

        contador = 2
        for i in self.features_length:
            new_contador = contador+i
            ws.merge_cells('A{}:B{}'.format(contador, new_contador-1))
            ws['L{}'.format(contador)] = '=SUM(K{}:K{})'.format(contador, new_contador-1)
            ws.merge_cells('L{}:L{}'.format(contador, new_contador-1))
            contador = new_contador

        for letra in abc: ws.column_dimensions[letra].width = 12.89
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['N'].width = 8
        ws.column_dimensions['K'].hidden= True

        ws['O3'].value = 'KS'
        ws['O4'].value = 'GINI'
        ws['P2'].value = 'Train'
        ws['Q2'].value = 'Test'
        ws['P3'].value = self.ks_train
        ws['Q3'].value = self.ks_test
        ws['P4'].value = self.gini_train
        ws['Q4'].value = self.gini_test

        for celda in ['P2', 'Q2', 'O3', 'O4']:
            cell_style(ws[celda], bold=True, hor_alignment='center', ver_alignment='center', 
            all_borders=True, font_color='ffffff', background_color='ff0000')

        for celda in ['P3', 'Q3', 'P4', 'Q4']:
            cell_style(ws[celda], hor_alignment='center', 
            ver_alignment='center', all_borders=True, wrap_text=True)
            ws[celda].number_format = '0.00%'
                   
        if color == 'green': color = 'CCFFCC'
        if color == 'light_blue': color = 'CCFFFF'
        if color == 'blue': color = 'CCECFF'
        if color == 'pink': color = 'FFCCFF'
        if color == 'red': color = 'FFCCCC'
        if color == 'yellow': color = 'FFFFCC'
        if color == 'purple': color = 'CCCCFE'
        if color == 'orange': color = 'FFCC99'
               
        contador, moneda = 2, 0
        for i in self.features_length:
            new_contador = contador+i
            if moneda%2 == 0:
                for row in ws['A{}:M{}'.format(contador, new_contador-1)]:
                    for cell in row:
                        cell.fill = PatternFill(fill_type='solid', fgColor=color)
            contador = new_contador
            moneda += 1
            
        wb.save(ruta)

    
    def create_pyspark_formula(self):
    
        import copy

        pyspark_formula = []

        for i in self.features:

            aux = 'CASE '
            points = list(self.scorecard[self.scorecard['Variable'] == i]['Aligned score'])
            groups = copy.deepcopy(list(self.scorecard[self.scorecard['Variable'] == i]['Group']))
            
            for j in range(len(groups)):

                if self.autogroupings[i].dtype not in ('object', 'bool'):
                    if 'Missing' in groups[j]:
                        aux += 'WHEN (isnan({}) OR ({} IS NULL)) THEN {} '.format(i, i, points[j])
                    if 'inf)' not in groups[j] and groups[j] != 'Missing':
                        lim = groups[j].split(', ')[1][:-1]
                        aux += 'WHEN {} < {} THEN {} '.format(i, lim, points[j])
                    if 'inf)' in groups[j]:
                        lim = groups[j].split(', ')[0][1:]
                        aux += 'WHEN {} >= {} THEN {} '.format(i, lim, points[j])

                else:
                    if 'Missing' in groups[j]:
                        todos = [aaa for bbb in groups for aaa in bbb] # CAMBIAR
                        if True in todos or False in todos: # CAMBIAR
                            aux += 'WHEN {} IS NULL THEN {} '.format(i, points[j]) # CAMBIAR
                        else: # CAMBIAR
                            aux += 'WHEN (isnan({}) OR ({} IS NULL)) THEN {} '\
                            .format(i, i, points[j]) # CAMBIAR
                    try: groups[j].remove('Missing')
                    except: pass
                    if groups[j] != []: 
                        aux += 'WHEN {} IN {} THEN {} '.format(i, groups[j], points[j])

            aux += 'ELSE {} END'.format(min(points))
            aux = aux.replace('[', '(').replace(']', ')')

            pyspark_formula.append(aux)
        
        self.pyspark_formula = pyspark_formula

        
    def create_pmml(self, nombre_archivo, 
    nombre_modelo='mew', score_name='scorecardpoints', reasons_code=False):

        import copy

        texto = '<PMML xmlns="http://www.dmg.org/PMML-4_4" version="4.4">\n'
        texto += '<DataDictionary>\n'

        for feature in self.features:
            objeto = self.autogroupings[feature]
            if 'int' in str(objeto.dtype): 
                tipo1 = 'integer'
                tipo2 = 'continuous'
            elif 'float' in str(objeto.dtype): 
                tipo1 = 'double'
                tipo2 = 'continuous'
            elif 'object' in str(objeto.dtype): 
                tipo1 = 'string'
                tipo2 = 'categorical'
            elif 'bool' in str(objeto.dtype): 
                tipo1 = 'boolean'
                tipo2 = 'categorical'
            else: 
                raise Exception('WTF, qué tipo de datos tiene {}: '\
                .format(feature), str(objeto.dtype))
            texto += '<DataField name="{}" dataType="{}" optype="{}"/>\n'\
            .format(feature, tipo1, tipo2)
        
        min_score = min(self.scorecard['Aligned score'])
        texto += '<DataField name="{}" dataType="integer" '.format(score_name) +\
        'optype="continuous"/>\n'
        texto += '</DataDictionary>\n'
        if reasons_code:
            texto += '<Scorecard modelName="{}" functionName="regression" '.format(nombre_modelo) +\
            'initialScore="0" useReasonCodes="true" reasonCodeAlgorithm="pointsAbove" ' +\
            'baselineScore="{}">\n'.format(min_score)
        else:
            texto += '<Scorecard modelName="{}" functionName="regression" '.format(nombre_modelo) +\
            'initialScore="0" useReasonCodes="false">\n'
        texto += '<MiningSchema>\n'

        for feature in self.features:
            texto += '<MiningField name="{}" usageType="active" '.format(feature) +\
            'invalidValueTreatment="asMissing"/>\n'

        texto += '<MiningField name="{}" usageType="predicted"/>\n'.format(score_name)
        texto += '</MiningSchema>\n'
        if reasons_code:
            texto += '<Output>\n'
            contador = 1
            for i in self.features:
                texto += '<OutputField name="reasoncode_{}" rank="{}"'.format(contador, contador) +\
                ' feature="reasonCode" dataType="string" optype="categorical"/>\n'
                contador += 1
            texto += '</Output>\n'     
        texto += '<Characteristics>\n'
        contador = 1
        for i in self.features:
            
            contador += 1
            aux = '<Characteristic name="{}_scr">\n'.format(i)
            points = list(self.scorecard[self.scorecard['Variable'] == i]['Aligned score'])
            groups = copy.deepcopy(list(self.scorecard[self.scorecard['Variable'] == i]['Group']))
            
            for j in range(len(points)):
                
                if reasons_code: 
                    aux += '<Attribute partialScore="{}" '.format(points[j]) +\
                    'reasonCode="{}_{}">\n'.format(i, j+1)
                else: aux += '<Attribute partialScore="{}">\n'.format(points[j])
                
                if self.autogroupings[i].dtype not in ('object', 'bool'):
  
                    if 'Missing' in groups[j]:
                        if len(groups[j]) > 7: aux += '<CompoundPredicate booleanOperator="or">\n'
                        aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                    if 'inf)' not in groups[j] and groups[j] != 'Missing':
                        lim = groups[j].split(', ')[1][:-1]
                        aux += '<SimplePredicate field="{}" operator="lessThan" '.format(i) +\
                        'value="{}"/>\n'.format(lim)
                    if 'inf)' in groups[j]:
                        lim = groups[j].split(', ')[0][1:]
                        aux += '<SimplePredicate field="{}" operator="greaterOrEqual" '.format(i) +\
                        'value="{}"/>\n'.format(lim)
                    if 'Missing' in groups[j] and len(groups[j]) > 7:
                        aux += '</CompoundPredicate>\n'
        
                else:
                
                    if len(groups[j]) > 1:
                        aux += '<CompoundPredicate booleanOperator="or">\n'
                    for k in groups[j]:
                        if k != 'Missing':
                            aux += '<SimplePredicate field="{}" operator="equal" '.format(i) +\
                            'value="{}"/>\n'.format(k)
                        else:
                            aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                    if len(groups[j]) > 1:
                        aux += '</CompoundPredicate>\n'
            
                aux += '</Attribute>\n'
            if reasons_code: 
                aux += '<Attribute partialScore="{}" '.format(min(points)) +\
                'reasonCode="{}_{}">\n'.format(i, 404)
            else: aux += '<Attribute partialScore="{}">\n'.format(min(points))
            aux += '<True/>\n'
            aux += '</Attribute>\n'
            aux += '</Characteristic>\n'
            texto += aux
        texto += '</Characteristics>\n'
        texto += '</Scorecard>\n'
        texto += '</PMML>'
        
        with open('{}'.format(nombre_archivo), 'w') as f: f.write(texto)


    def create_pmml_str(self, nombre_modelo='mew', score_name='scorecardpoints', reasons_code=False):

            import copy

            texto = '<PMML xmlns="http://www.dmg.org/PMML-4_4" version="4.4">\n'
            texto += '<DataDictionary>\n'

            for feature in self.features:
                objeto = self.autogroupings[feature]
                if 'int' in str(objeto.dtype): 
                    tipo1 = 'integer'
                    tipo2 = 'continuous'
                elif 'float' in str(objeto.dtype): 
                    tipo1 = 'double'
                    tipo2 = 'continuous'
                elif 'object' in str(objeto.dtype): 
                    tipo1 = 'string'
                    tipo2 = 'categorical'
                elif 'bool' in str(objeto.dtype): 
                    tipo1 = 'boolean'
                    tipo2 = 'categorical'
                else: 
                    raise Exception('WTF, qué tipo de datos tiene {}: '\
                    .format(feature), str(objeto.dtype))
                texto += '<DataField name="{}" dataType="{}" optype="{}"/>\n'\
                .format(feature, tipo1, tipo2)
            
            min_score = min(self.scorecard['Aligned score'])
            texto += '<DataField name="{}" dataType="integer" '.format(score_name) +\
            'optype="continuous"/>\n'
            texto += '</DataDictionary>\n'
            if reasons_code:
                texto += '<Scorecard modelName="{}" functionName="regression" '.format(nombre_modelo) +\
                'initialScore="0" useReasonCodes="true" reasonCodeAlgorithm="pointsAbove" ' +\
                'baselineScore="{}">\n'.format(min_score)
            else:
                texto += '<Scorecard modelName="{}" functionName="regression" '.format(nombre_modelo) +\
                'initialScore="0" useReasonCodes="false">\n'
            texto += '<MiningSchema>\n'

            for feature in self.features:
                texto += '<MiningField name="{}" usageType="active" '.format(feature) +\
                'invalidValueTreatment="asMissing"/>\n'

            texto += '<MiningField name="{}" usageType="predicted"/>\n'.format(score_name)
            texto += '</MiningSchema>\n'
            if reasons_code:
                texto += '<Output>\n'
                contador = 1
                for i in self.features:
                    texto += '<OutputField name="reasoncode_{}" rank="{}"'.format(contador, contador) +\
                    ' feature="reasonCode" dataType="string" optype="categorical"/>\n'
                    contador += 1
                texto += '</Output>\n'     
            texto += '<Characteristics>\n'
            contador = 1
            for i in self.features:
                
                contador += 1
                aux = '<Characteristic name="{}_scr">\n'.format(i)
                points = list(self.scorecard[self.scorecard['Variable'] == i]['Aligned score'])
                groups = copy.deepcopy(list(self.scorecard[self.scorecard['Variable'] == i]['Group']))
                
                for j in range(len(points)):
                    
                    if reasons_code: 
                        aux += '<Attribute partialScore="{}" '.format(points[j]) +\
                        'reasonCode="{}_{}">\n'.format(i, j+1)
                    else: aux += '<Attribute partialScore="{}">\n'.format(points[j])
                    
                    if self.autogroupings[i].dtype not in ('object', 'bool'):
    
                        if 'Missing' in groups[j]:
                            if len(groups[j]) > 7: aux += '<CompoundPredicate booleanOperator="or">\n'
                            aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                        if 'inf)' not in groups[j] and groups[j] != 'Missing':
                            lim = groups[j].split(', ')[1][:-1]
                            aux += '<SimplePredicate field="{}" operator="lessThan" '.format(i) +\
                            'value="{}"/>\n'.format(lim)
                        if 'inf)' in groups[j]:
                            lim = groups[j].split(', ')[0][1:]
                            aux += '<SimplePredicate field="{}" operator="greaterOrEqual" '.format(i) +\
                            'value="{}"/>\n'.format(lim)
                        if 'Missing' in groups[j] and len(groups[j]) > 7:
                            aux += '</CompoundPredicate>\n'
            
                    else:
                    
                        if len(groups[j]) > 1:
                            aux += '<CompoundPredicate booleanOperator="or">\n'
                        for k in groups[j]:
                            if k != 'Missing':
                                aux += '<SimplePredicate field="{}" operator="equal" '.format(i) +\
                                'value="{}"/>\n'.format(k)
                            else:
                                aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                        if len(groups[j]) > 1:
                            aux += '</CompoundPredicate>\n'
                
                    aux += '</Attribute>\n'
                if reasons_code: 
                    aux += '<Attribute partialScore="{}" '.format(min(points)) +\
                    'reasonCode="{}_{}">\n'.format(i, 404)
                else: aux += '<Attribute partialScore="{}">\n'.format(min(points))
                aux += '<True/>\n'
                aux += '</Attribute>\n'
                aux += '</Characteristic>\n'
                texto += aux
            texto += '</Characteristics>\n'
            texto += '</Scorecard>\n'
            texto += '</PMML>'
            
            return texto  
        
    def predict(self, data, target_name='', keep_columns=[], binary_treshold=0.0):
                   
        if target_name != '': X1 = data[keep_columns + self.features + [target_name]].copy()
        else: X1 = data[keep_columns + self.features].copy()

        X1_v2, info = X1.copy(), {}
        for feature in self.features:

            info[feature] = {}

            bp = self.final_breakpoints[feature]

            if not isinstance(bp, dict):
                X1_v2[feature] = data_convert(X1[feature].values, string_categories2(bp))[3]
                info[feature]['breakpoints_num'] = breakpoints_to_num(bp)
                info[feature]['group_names'] = compute_group_names(X1[feature].values.dtype, bp)

            else:
                X1_v2[feature] = remapeo_missing(data_convert(
                X1[feature].values, string_categories2(bp))[3], bp)
                info[feature]['breakpoints_num'] = breakpoints_to_num(bp['bp'])
                info[feature]['group_names'] = compute_group_names(
                X1[feature].values.dtype, bp['bp'], bp['mg'])
                
        X1_v2 = X1_v2.reset_index(drop=True)
        X2 = apply_scorecard(X1_v2, self.scorecard, info, 
        target_name=target_name, binary_treshold=binary_treshold)

        venga = 0
        for i in X2.columns:
            if 'scr_' in i:
                break
            venga += 1

        for i in X2.columns[venga:]: X1[i] = X2[i]

        return X1
        

class Autogrouping:


    def __init__(self, max_groups=5, min_pct=0.05, manual_type=''):

        self.max_groups = max_groups
        self.min_pct = min_pct
        self.manual_type = manual_type


    def fit(self, x, y):
        
        if self.manual_type != '': dtype = self.manual_type
        else: dtype = x.dtype

        if dtype not in ('O', 'bool'): categories = {}

        else:

            categories = string_categories1(x, y)

            if pd.Series(x).isna().sum() > 0:

                for i in categories:
                    if not isinstance(i, str) and not isinstance(i, bool):
                        aux_miss = i

                categories['Missing'] = categories.pop(aux_miss) 
                categories = dict(sorted(categories.items(), key=lambda item: item[1]))

        frenken = data_convert(x, categories)
        x_converted = frenken[2]
        self.x_final = frenken[3]

        if dtype not in ('O', 'bool') and np.isnan(x_converted).sum() > 0:

            aux = ~np.isnan(x_converted)
            x_nm, y_nm = x_converted[aux], y[aux]

        else: x_nm, y_nm = x_converted, y

        self.compute_groups(x_nm, y_nm)

        missing_group = None

        if dtype not in ('O', 'bool') and np.isnan(x_converted).sum() > 0:

            self.breakpoints_num = np.array([-12345670] + list(self.breakpoints_num))

            x_groups = np.digitize(self.x_final, self.breakpoints_num)

            ngroups = len(self.breakpoints_num) + 1
            g = np.zeros(ngroups).astype(np.int64)
            b = np.zeros(ngroups).astype(np.int64)

            for i in range(ngroups):
                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])

            missing_group = 0

            if b[0] == 0 or g[0] == 0:

                self.breakpoints_num = self.breakpoints_num[1:]

                keo = b[1:]/(b[1:]+g[1:])
                if b[0] == 0: indice = int(np.where(keo == keo.min())[0][0])
                if g[0] == 0: indice = int(np.where(keo == keo.max())[0][0])

                if indice == 0:
                    self.x_final[self.x_final == -12345678] = \
                    self.breakpoints_num[indice] - (np.e-2)

                else:
                    self.x_final[self.x_final == -12345678] = \
                    self.breakpoints_num[indice-1] + (np.e-2)

                missing_group = indice + 1

        if dtype in ('O', 'bool'): 
            self.breakpoints = breakpoints_to_str(self.breakpoints_num, categories)
        else: self.breakpoints = {'bp': self.breakpoints_num, 'mg': missing_group}

        group_names = compute_group_names(dtype, self.breakpoints, missing_group)
        self.table, self.iv = compute_table(self.x_final, y, self.breakpoints_num, group_names)
            
        self.missing = not missing_group == None
        self.categories = categories
        self.dtype = dtype

        return self


    def compute_groups(self, x, y): 

        tree_args = {
            'max_leaf_nodes': self.max_groups,
            'min_samples_leaf': self.min_pct
        }
        
        tree = DecisionTreeClassifier(**tree_args).fit(x.reshape(-1, 1), y)
        aux = np.unique(tree.tree_.threshold)
        breakpoints_num = aux[aux != _tree.TREE_UNDEFINED]

        x_groups = np.digitize(x, breakpoints_num)

        ngroups = len(breakpoints_num) + 1
        g = np.zeros(ngroups).astype(np.int64)
        b = np.zeros(ngroups).astype(np.int64)

        for i in range(ngroups):

            g[i] = np.sum([(y == 0) & (x_groups == i)])
            b[i] = np.sum([(y == 1) & (x_groups == i)])

        error = (g == 0) | (b == 0)

        while np.any(error):

            m_bk = np.concatenate([error[:-2], [error[-2] | error[-1]]])

            breakpoints_num = breakpoints_num[~m_bk]
            x_groups = np.digitize(x, breakpoints_num)

            ngroups = len(breakpoints_num) + 1
            g = np.zeros(ngroups).astype(np.int64)
            b = np.zeros(ngroups).astype(np.int64)

            for i in range(ngroups):

                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])

            error = (g == 0) | (b == 0)

        self.breakpoints_num = np.round(breakpoints_num, 4)
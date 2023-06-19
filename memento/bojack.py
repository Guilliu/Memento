import numpy as np, pandas as pd

from sklearn.model_selection import train_test_split
from sklearn.tree import _tree, DecisionTreeClassifier

from .todd import *
from .diane import *
from .mr_peanutbutter import *
from .princess_carolyn import *

####################################################################################################


class Scorecard:


    def __init__(
        
        self,
        
        test_seed=123,
        test_size=0.3,
        stratify=True,
        stratify_var='',
        flag_train_test=[],

        autogrp_all=False,
        autogrp_max_groups=5,
        autogrp_min_pct=0.05,
        autogrp_dict_max_groups={},
        autogrp_dict_min_pct={},
        autogrp_dict_manual_types={},

        features=[],
        excluded_vars=[],
        included_vars=[],

        iv_threshold=0.015,
        selection_show = 'gini',
        selection_method='stepwise',
        selection_metric='pvalue',
        selection_threshold=0.01,
        selection_max_iters=14,
        selection_stop_ks_gini=True,
        selection_check_overfitting=True,
        
        user_breakpoints={},
        logistic_method = 'newton',

        id_columns=[],
        save_tables='features',
        save_autogroupings='features'
        
    ):

        self.test_seed = test_seed
        self.test_size = test_size
        self.stratify = stratify
        self.stratify_var = stratify_var
        self.flag_train_test = flag_train_test
        
        self.autogrp_all = autogrp_all
        self.autogrp_max_groups = autogrp_max_groups
        self.autogrp_min_pct = autogrp_min_pct
        self.autogrp_dict_max_groups = autogrp_dict_max_groups
        self.autogrp_dict_min_pct = autogrp_dict_min_pct
        self.autogrp_dict_manual_types = autogrp_dict_manual_types
        
        self.features = features
        self.excluded_vars = excluded_vars
        self.included_vars = included_vars
        
        self.iv_threshold = iv_threshold
        self.selection_show = selection_show
        self.selection_method = selection_method
        self.selection_metric = selection_metric
        self.selection_threshold = selection_threshold
        self.selection_max_iters = selection_max_iters
        self.selection_stop_ks_gini = selection_stop_ks_gini
        self.selection_check_overfitting = selection_check_overfitting

        self.user_breakpoints = user_breakpoints
        self.logistic_method = logistic_method
        
        self.id_columns = id_columns
        self.save_tables = save_tables
        self.save_autogroupings = save_autogroupings
        

    def fit(self, X, y):

        ### INICIO PARTICIONADO

        if self.flag_train_test != []:

            try: a, b, c = self.flag_train_test
            except:
                print('En la variable flag_train_test hay que introducir tres cosas en orden: '
                'el nombre de la variable con el flag, el valor de train y el valor de test')

            data = X.copy()
            data['target_4815162342'] = y

            X_train = data[data[a] == b].drop('target_4815162342', axis=1)
            y_train = data[data[a] == b]['target_4815162342'].values

            X_test = data[data[a] == c].drop('target_4815162342', axis=1)
            y_test = data[data[a] == c]['target_4815162342'].values

        else:

            if self.stratify:

                if self.stratify_var == '':
                    X_train, X_test, y_train, y_test = train_test_split(X, y,
                    test_size=self.test_size, random_state=self.test_seed, stratify=y)
                    
                else:
                    X_train, X_test, y_train, y_test = train_test_split(X, y,
                    test_size=self.test_size, random_state=self.test_seed,
                    stratify=X[self.stratify_var])

            else:
                X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=self.test_size, random_state=self.test_seed)

        self.index_train, self.index_test = X_train.index, X_test.index
        X_train, X_test = X_train.reset_index(drop=True), X_test.reset_index(drop=True)
        
        if isinstance(y, pd.Series): y_train, y_test = y_train.values, y_test.values
        self.y_train, self.y_test = y_train, y_test
        
        if self.flag_train_test == []:

            if self.stratify:

                if self.stratify_var == '':
                    print('Particionado {}-{} estratificado en el target terminado'\
                    .format(int(100*(1-self.test_size)), int(100*self.test_size)))

                else:
                    print('Particionado {}-{} estratificado en la variable \'{}\' terminado'\
                    .format(int(100*(1-self.test_size)), 
                    int(100*self.test_size), self.stratify_var))

            else:
                print('Particionado {}-{} terminado'\
                .format(int(100*(1-self.test_size)), int(100*self.test_size)))
                
            print('-' * N)

        ### INICIO AUTOGROUPING
                
        if self.autogrp_all: variables = X.columns
        else: 
            if self.features != []: variables = self.features
            else: variables = list(set(list(X.columns)) - set(self.excluded_vars))

        autogroupings = {}
        variables_no_agrupadas_error = []

        for variable in variables:

            try:
                
                if variable in self.autogrp_dict_max_groups:
                    max_groups = self.autogrp_dict_max_groups[variable]
                else: max_groups = self.autogrp_max_groups

                if variable in self.autogrp_dict_min_pct:
                    min_pct = self.autogrp_dict_min_pct[variable]
                else: min_pct = self.autogrp_min_pct

                if variable in self.autogrp_dict_manual_types: 
                    manual_type = self.autogrp_dict_manual_types[variable]
                else: manual_type = ''

                x = X_train[variable].values
                frenken = Autogrouping(max_groups=max_groups, 
                min_pct=min_pct, manual_type=manual_type).fit(x, y_train)

                if len(frenken.breakpoints_num) == 0: variables_no_agrupadas_error.append(variable)
                else: autogroupings[variable] = frenken
                
            except:
                variables_no_agrupadas_error.append(variable)

        print('Autogrouping terminado. Máximo número de buckets = {}. Mínimo porcentaje '
        'por bucket = {}'.format(self.autogrp_max_groups, self.autogrp_min_pct))
        print('-' * N)

        if len(variables_no_agrupadas_error) > 0:
            print('Variables no agrupadas por algún error, seguramente por excesiva '
            'concentración en algún valor (> 95%) : {}'.format(variables_no_agrupadas_error))
            print('-' * N)
            
        self.variables_no_agrupadas_error = variables_no_agrupadas_error

        tabla_ivs, contador = pd.DataFrame(columns=['variable', 'iv']), 0
        for variable in autogroupings:
            tabla_ivs.loc[contador] = variable, autogroupings[variable].iv
            contador += 1
        tabla_ivs = tabla_ivs.sort_values('iv', ascending=False).reset_index(drop=True)

        self.tabla_ivs = tabla_ivs
        
        variables_filtro_iv = tabla_ivs[tabla_ivs['iv'] >= self.iv_threshold]['variable']
        
        if self.features != []: variables_def = self.features
        else:
            variables_def = list(set(variables_filtro_iv) - set(variables_no_agrupadas_error))

        self.final_breakpoints = compute_final_breakpoints(
        variables_def, autogroupings, self.user_breakpoints)
        info = compute_info(X_train, variables_def, self.final_breakpoints)

        df_train = adapt_data(X_train, y_train, variables_def, self.final_breakpoints)
        df_test = adapt_data(X_test, y_test, variables_def, self.final_breakpoints)

        features = self.features

        features = features_selection(
        df_train, self.features, variables_def, info, 'target_4815162342',
        method=self.selection_method, metric=self.selection_metric,
        threshold=self.selection_threshold, stop_ks_gini=self.selection_stop_ks_gini,
        max_iters=self.selection_max_iters, included_vars=self.included_vars,
        muestra_test=df_test, show=self.selection_show, 
        check_overfitting=self.selection_check_overfitting,
        logistic_method=self.logistic_method)

        df_train = df_train[features + ['target_4815162342']]
        df_test = df_test[features + ['target_4815162342']]

        scorecard, features_length, pvalues, coefs = compute_scorecard(
        df_train, features, info, pvalues=True, 
        ret_coefs=True, logistic_method=self.logistic_method)

        df_train_final = apply_scorecard(df_train, scorecard, info, 'target_4815162342')
        ks_train, gini_train = compute_metrics(df_train_final, 'target_4815162342', ['gini', 'ks'])

        print('El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de entrenamiento'.format(round(ks_train*100, 2), round(gini_train*100, 2)))
        print('-' * N)

        df_test_final = apply_scorecard(df_test, scorecard, info, 'target_4815162342')
        ks_test, gini_test = compute_metrics(df_test_final, 'target_4815162342', ['gini', 'ks']) 

        print('El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de validación'.format(round(ks_test*100, 2), round(gini_test*100, 2)))
        print('-' * N)

        self.features = features
        self.scorecard = scorecard
        self.features_length = features_length
        self.pvalues = dict(zip(features, list(pvalues[1:])))
        self.coefs = coefs

        self.ks_train = ks_train
        self.gini_train = gini_train
        self.ks_test = ks_test
        self.gini_test = gini_test

        if self.save_tables == 'all': 
            self.X_train = X_train
            self.X_test = X_test
        elif self.save_tables == 'features':
            self.X_train = X_train[self.id_columns + features]
            self.X_test = X_test[self.id_columns + features]

        for objeto in autogroupings: del autogroupings[objeto].x_final

        if self.save_autogroupings == 'all': 
            self.autogroupings = autogroupings
        elif self.save_autogroupings == 'features':
            self.autogroupings = dict((k, autogroupings[k]) for k in features if k in autogroupings)
            
        try: self.create_pyspark_formula()
        except: self.pypsark_formula = ['Por algún motivo no se ha podido calcular la fórmula SQL']

        return self

        
    def create_excel(self, ruta, color='blue'):
        
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill
        
        scorecard = self.scorecard.copy()
        scorecard = scorecard.drop('Raw score', axis=1)
        
        abc = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
        'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
        'AD', 'AE', 'AF', 'AG','AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP',
        'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        rows = dataframe_to_rows(scorecard, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                try: ws.cell(row=r_idx, column=c_idx, value=value)
                except: ws.cell(row=r_idx, column=c_idx, value=str(value))

        ws.insert_cols(2)
        ws.insert_cols(4)
        ws.insert_cols(12)

        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:D1')

        altura = len(scorecard)

        for letra in ['F', 'I']:
            for row in ws['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.00%'

        for letra in ['J', 'K', 'L']:
            for row in ws['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.0000'

        for i in range(2, altura+2):
            ws.merge_cells('C{}:D{}'.format(i, i))

        for row in ws['A1:M1']:
            for cell in row:
                cell_style(cell, bold=True, hor_alignment='center', ver_alignment='center', 
                all_borders=True, font_color='ffffff', background_color='ff0000')

        for row in ws['A2:M{}'.format(altura+1)]:
            for cell in row:
                cell_style(cell, hor_alignment='center', 
                ver_alignment='center', all_borders=True, wrap_text=True)

        ws['K1'].value = 'IV aux'
        ws['L1'].value = 'IV'

        contador = 2
        for i in self.features_length:
            new_contador = contador+i
            ws.merge_cells('A{}:B{}'.format(contador, new_contador-1))
            ws['L{}'.format(contador)] = '=SUM(K{}:K{})'.format(contador, new_contador-1)
            ws.merge_cells('L{}:L{}'.format(contador, new_contador-1))
            contador = new_contador

        for letra in abc: ws.column_dimensions[letra].width = 12.89
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['N'].width = 8
        ws.column_dimensions['K'].hidden= True

        ws['O3'].value = 'KS'
        ws['O4'].value = 'GINI'
        ws['P2'].value = 'Train'
        ws['Q2'].value = 'Test'
        ws['P3'].value = self.ks_train
        ws['Q3'].value = self.ks_test
        ws['P4'].value = self.gini_train
        ws['Q4'].value = self.gini_test

        for celda in ['P2', 'Q2', 'O3', 'O4']:
            cell_style(ws[celda], bold=True, hor_alignment='center', ver_alignment='center', 
            all_borders=True, font_color='ffffff', background_color='ff0000')

        for celda in ['P3', 'Q3', 'P4', 'Q4']:
            cell_style(ws[celda], hor_alignment='center', 
            ver_alignment='center', all_borders=True, wrap_text=True)
            ws[celda].number_format = '0.00%'
                   
        if color == 'green': color = 'CCFFCC'
        if color == 'light_blue': color = 'CCFFFF'
        if color == 'blue': color = 'CCECFF'
        if color == 'pink': color = 'FFCCFF'
        if color == 'red': color = 'FFCCCC'
        if color == 'yellow': color = 'FFFFCC'
        if color == 'purple': color = 'CCCCFE'
        if color == 'orange': color = 'FFCC99'
               
        contador, moneda = 2, 0
        for i in self.features_length:
            new_contador = contador+i
            if moneda%2 == 0:
                for row in ws['A{}:M{}'.format(contador, new_contador-1)]:
                    for cell in row:
                        cell.fill = PatternFill(fill_type='solid', fgColor=color)
            contador = new_contador
            moneda += 1
            
        wb.save(ruta)

    
    def create_pyspark_formula(self):
    
        import copy

        pyspark_formula = []

        for i in self.features:

            aux = 'CASE '
            points = list(self.scorecard[self.scorecard['Variable'] == i]['Aligned score'])
            groups = copy.deepcopy(list(self.scorecard[self.scorecard['Variable'] == i]['Group']))
            
            for j in range(len(groups)):

                if self.autogroupings[i].dtype not in ('object', 'bool'):
                    if 'Missing' in groups[j]:
                        aux += 'WHEN (isnan({}) OR ({} IS NULL)) THEN {} '.format(i, i, points[j])
                    if 'inf)' not in groups[j] and groups[j] != 'Missing':
                        lim = groups[j].split(', ')[1][:-1]
                        aux += 'WHEN {} < {} THEN {} '.format(i, lim, points[j])
                    if 'inf)' in groups[j]:
                        lim = groups[j].split(', ')[0][1:]
                        aux += 'WHEN {} >= {} THEN {} '.format(i, lim, points[j])

                else:
                    if 'Missing' in groups[j]:
                        todos = [aaa for bbb in groups for aaa in bbb] # CAMBIAR
                        if True in todos or False in todos: # CAMBIAR
                            aux += 'WHEN {} IS NULL THEN {} '.format(i, points[j]) # CAMBIAR
                        else: # CAMBIAR
                            aux += 'WHEN (isnan({}) OR ({} IS NULL)) THEN {} '\
                            .format(i, i, points[j]) # CAMBIAR
                    try: groups[j].remove('Missing')
                    except: pass
                    if groups[j] != []: 
                        aux += 'WHEN {} IN {} THEN {} '.format(i, groups[j], points[j])

            aux += 'ELSE {} END'.format(min(points))
            aux = aux.replace('[', '(').replace(']', ')')

            pyspark_formula.append(aux)
        
        self.pyspark_formula = pyspark_formula

        
    def create_pmml(self, nombre_archivo, 
    nombre_modelo='mew', score_name='scorecardpoints', reasons_code=True):

        import copy

        texto = '<PMML xmlns="http://www.dmg.org/PMML-4_4" version="4.4">\n'
        texto += '<DataDictionary>\n'

        for feature in self.features:
            objeto = self.autogroupings[feature]
            if 'int' in str(objeto.dtype): 
                tipo1 = 'integer'
                tipo2 = 'continuous'
            elif 'float' in str(objeto.dtype): 
                tipo1 = 'double'
                tipo2 = 'continuous'
            elif 'object' in str(objeto.dtype): 
                tipo1 = 'string'
                tipo2 = 'categorical'
            elif 'bool' in str(objeto.dtype): 
                tipo1 = 'boolean'
                tipo2 = 'categorical'
            else: 
                raise Exception('WTF, qué tipo de datos tiene {}: '\
                .format(feature), str(objeto.dtype))
            texto += '<DataField name="{}" dataType="{}" optype="{}"/>\n'\
            .format(feature, tipo1, tipo2)
        
        min_score = min(self.scorecard['Aligned score'])
        texto += '<DataField name="{}" dataType="integer" ' +\
        'optype="continuous"/>\n'.format(score_name)
        texto += '</DataDictionary>\n'
        if reasons_code:
            texto += '<Scorecard modelName="{}" functionName="regression" initialScore="0" ' +\
            'useReasonCodes="true" reasonCodeAlgorithm="pointsAbove" ' +\
            'baselineScore="{}">\n'.format(nombre_modelo, min_score)
        else:
            texto += '<Scorecard modelName="{}" functionName="regression" ' +\
            'initialScore="0" useReasonCodes="false">\n'.format(nombre_modelo, min_score)
        texto += '<MiningSchema>\n'

        for feature in self.features:
            texto += '<MiningField name="{}" usageType="active" ' +\
            'invalidValueTreatment="asMissing"/>\n'.format(feature)

        texto += '<MiningField name="{}" usageType="predicted"/>\n'.format(score_name)
        texto += '</MiningSchema>\n'
        if reasons_code:
            texto += '<Output>\n'
            contador = 1
            for i in self.features:
                texto += '<OutputField name="reasoncode_{}" rank="{}" feature="reasonCode" ' +\
                'dataType="string" optype="categorical"/>\n'.format(contador, contador)
                contador += 1
            texto += '</Output>\n'     
        texto += '<Characteristics>\n'
        contador = 1
        for i in self.features:
            
            contador += 1
            aux = '<Characteristic name="{}">\n'.format(i)
            points = list(self.scorecard[self.scorecard['Variable'] == i]['Aligned score'])
            groups = copy.deepcopy(list(self.scorecard[self.scorecard['Variable'] == i]['Group']))
            
            for j in range(len(points)):
                
                if reasons_code: 
                    aux += '<Attribute partialScore="{}" ' +\
                    'reasonCode="{}_{}">\n'.format(points[j], i, j+1)
                else: aux += '<Attribute partialScore="{}">\n'.format(points[j])
                
                if self.autogroupings[i].dtype not in ('object', 'bool'):
  
                    if 'Missing' in groups[j]:
                        if len(groups[j]) > 7: aux += '<CompoundPredicate booleanOperator="or">\n'
                        aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                    if 'inf)' not in groups[j] and groups[j] != 'Missing':
                        lim = groups[j].split(', ')[1][:-1]
                        aux += '<SimplePredicate field="{}" operator="lessThan" ' +\
                        'value="{}"/>\n'.format(i, lim)
                    if 'inf)' in groups[j]:
                        lim = groups[j].split(', ')[0][1:]
                        aux += '<SimplePredicate field="{}" operator="greaterOrEqual" ' +\
                        'value="{}"/>\n'.format(i, lim)
                    if 'Missing' in groups[j] and len(groups[j]) > 7:
                        aux += '</CompoundPredicate>\n'
        
                else:
                
                    if len(groups[j]) > 1:
                        aux += '<CompoundPredicate booleanOperator="or">\n'
                    for k in groups[j]:
                        if k != 'Missing':
                            aux += '<SimplePredicate field="{}" operator="equal" ' +\
                            'value="{}"/>\n'.format(i, k)
                        else:
                            aux += '<SimplePredicate field="{}" operator="isMissing"/>\n'.format(i)
                    if len(groups[j]) > 1:
                        aux += '</CompoundPredicate>\n'
            
                aux += '</Attribute>\n'
            if reasons_code: 
                aux += '<Attribute partialScore="{}" ' +\
                'reasonCode="{}_{}">\n'.format(min(points), i, 404)
            else: aux += '<Attribute partialScore="{}">\n'.format(min(points))
            aux += '<True/>\n'
            aux += '</Attribute>\n'
            aux += '</Characteristic>\n'
            texto += aux
        texto += '</Characteristics>\n'
        texto += '</Scorecard>\n'
        texto += '</PMML>'
        
        with open('{}'.format(nombre_archivo), 'w') as f: f.write(texto)
    
        
    def predict(self, data, target_name='', keep_columns=[], binary_treshold=0.0):
                   
        if target_name != '': X1 = data[keep_columns + self.features + [target_name]].copy()
        else: X1 = data[keep_columns + self.features].copy()

        X1_v2, info = X1.copy(), {}
        for feature in self.features:

            info[feature] = {}

            bp = self.final_breakpoints[feature]

            if not isinstance(bp, dict):
                X1_v2[feature] = data_convert(X1[feature].values, string_categories2(bp))[3]
                info[feature]['breakpoints_num'] = breakpoints_to_num(bp)
                info[feature]['group_names'] = compute_group_names(X1[feature].values.dtype, bp)

            else:
                X1_v2[feature] = remapeo_missing(data_convert(
                X1[feature].values, string_categories2(bp))[3], bp)
                info[feature]['breakpoints_num'] = breakpoints_to_num(bp['bp'])
                info[feature]['group_names'] = compute_group_names(
                X1[feature].values.dtype, bp['bp'], bp['mg'])
                
        X1_v2 = X1_v2.reset_index(drop=True)
        X2 = apply_scorecard(X1_v2, self.scorecard, info, 
        target_name=target_name, binary_treshold=binary_treshold)

        venga = 0
        for i in X2.columns:
            if 'scr_' in i:
                break
            venga += 1

        for i in X2.columns[venga:]: X1[i] = X2[i]

        return X1
        

class Autogrouping:


    def __init__(self, max_groups=5, min_pct=0.05, manual_type=''):

        self.max_groups = max_groups
        self.min_pct = min_pct
        self.manual_type = manual_type


    def fit(self, x, y):
        
        if self.manual_type != '': dtype = self.manual_type
        else: dtype = x.dtype

        if dtype not in ('O', 'bool'): categories = {}

        else:

            categories = string_categories1(x, y)

            if pd.Series(x).isna().sum() > 0:

                for i in categories:
                    if not isinstance(i, str) and not isinstance(i, bool):
                        aux_miss = i

                categories['Missing'] = categories.pop(aux_miss) 
                categories = dict(sorted(categories.items(), key=lambda item: item[1]))

        frenken = data_convert(x, categories)
        x_converted = frenken[2]
        self.x_final = frenken[3]

        if dtype not in ('O', 'bool') and np.isnan(x_converted).sum() > 0:

            aux = ~np.isnan(x_converted)
            x_nm, y_nm = x_converted[aux], y[aux]

        else: x_nm, y_nm = x_converted, y

        self.compute_groups(x_nm, y_nm)

        missing_group = None

        if dtype not in ('O', 'bool') and np.isnan(x_converted).sum() > 0:

            self.breakpoints_num = np.array([-12345670] + list(self.breakpoints_num))

            x_groups = np.digitize(self.x_final, self.breakpoints_num)

            ngroups = len(self.breakpoints_num) + 1
            g = np.zeros(ngroups).astype(np.int64)
            b = np.zeros(ngroups).astype(np.int64)

            for i in range(ngroups):
                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])

            missing_group = 0

            if b[0] == 0 or g[0] == 0:

                self.breakpoints_num = self.breakpoints_num[1:]

                keo = b[1:]/(b[1:]+g[1:])
                if b[0] == 0: indice = int(np.where(keo == keo.min())[0][0])
                if g[0] == 0: indice = int(np.where(keo == keo.max())[0][0])

                if indice == 0:
                    self.x_final[self.x_final == -12345678] = \
                    self.breakpoints_num[indice] - (np.e-2)

                else:
                    self.x_final[self.x_final == -12345678] = \
                    self.breakpoints_num[indice-1] + (np.e-2)

                missing_group = indice + 1

        if dtype in ('O', 'bool'): 
            breakpoints = breakpoints_to_str(self.breakpoints_num, categories)
        else: breakpoints = self.breakpoints_num

        group_names = compute_group_names(dtype, breakpoints, missing_group)
        self.table, self.iv = compute_table(self.x_final, y, self.breakpoints_num, group_names)

        self.breakpoints = breakpoints
        self.missing_group = missing_group
        self.categories = categories
        self.dtype = dtype

        return self


    def compute_groups(self, x, y): 

        tree_args = {
            'max_leaf_nodes': self.max_groups,
            'min_samples_leaf': self.min_pct
        }
        
        tree = DecisionTreeClassifier(**tree_args).fit(x.reshape(-1, 1), y)
        aux = np.unique(tree.tree_.threshold)
        breakpoints_num = aux[aux != _tree.TREE_UNDEFINED]

        x_groups = np.digitize(x, breakpoints_num)

        ngroups = len(breakpoints_num) + 1
        g = np.zeros(ngroups).astype(np.int64)
        b = np.zeros(ngroups).astype(np.int64)

        for i in range(ngroups):

            g[i] = np.sum([(y == 0) & (x_groups == i)])
            b[i] = np.sum([(y == 1) & (x_groups == i)])

        error = (g == 0) | (b == 0)

        while np.any(error):

            m_bk = np.concatenate([error[:-2], [error[-2] | error[-1]]])

            breakpoints_num = breakpoints_num[~m_bk]
            x_groups = np.digitize(x, breakpoints_num)

            ngroups = len(breakpoints_num) + 1
            g = np.zeros(ngroups).astype(np.int64)
            b = np.zeros(ngroups).astype(np.int64)

            for i in range(ngroups):

                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])

            error = (g == 0) | (b == 0)

        self.breakpoints_num = breakpoints_num

